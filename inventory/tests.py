from datetime import datetime, timedelta
import io
import json
import re
import tempfile
from pathlib import Path
from unittest.mock import patch

from django.conf import settings
from django.core import mail
from django.contrib.auth.models import User
from django.test import Client, TestCase, override_settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from PIL import Image

from .middleware import LAST_ACTIVITY_KEY
from .models import (
    DropdownOption,
    InventorySection,
    MutcdClassification,
    MutcdFallback,
    MutcdMapping,
    Project,
    RegistrationApproval,
    TabRecord,
)


class InventoryImageTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("image-owner", password="StrongPass!234")
        self.other = User.objects.create_user("other-user", password="StrongPass!234")
        self.project = Project.objects.create(name="Image Project", code="image-project")
        self.project.members.add(self.user)
        other_project = Project.objects.create(name="Other Project", code="other-project")
        other_project.members.add(self.other)
        self.record = TabRecord.objects.create(
            project=self.project, owner=self.user, tab="sign", tab_record_id=1, data={"IMAGE_LINK": ""}
        )

    @staticmethod
    def image_upload(image_format, color):
        stream = io.BytesIO()
        Image.new("RGB", (12, 8), color).save(stream, format=image_format)
        extension = "jpg" if image_format == "JPEG" else image_format.lower()
        return SimpleUploadedFile(
            f"test.{extension}", stream.getvalue(), content_type=f"image/{extension}"
        )

    def test_upload_sets_full_url_and_reupload_replaces_existing_image(self):
        with tempfile.TemporaryDirectory() as temporary_root, override_settings(
            INVENTORY_UPLOAD_ROOT=Path(temporary_root), APP_BASE_URL="http://testserver"
        ):
            self.client.force_login(self.user)
            url = reverse("api_record_image", args=["sign", 1])
            first = self.client.post(url, {"image": self.image_upload("JPEG", "red")})
            self.assertEqual(first.status_code, 200)
            first_link = first.json()["record"]["IMAGE_LINK"]
            self.assertTrue(first_link.startswith("http://testserver/uploads/images/sign_inventory/1/"))
            directory = Path(temporary_root) / "images" / "sign_inventory" / "1"
            self.assertEqual([path.name for path in directory.iterdir()], ["image.jpg"])

            second = self.client.post(url, {"image": self.image_upload("PNG", "blue")})
            self.assertEqual(second.status_code, 200)
            self.assertEqual([path.name for path in directory.iterdir()], ["image.png"])
            self.record.refresh_from_db()
            self.assertTrue(self.record.data["IMAGE_LINK"].endswith("/image.png"))
            view = self.client.get(self.record.data["IMAGE_LINK"])
            self.assertEqual(view.status_code, 200)
            self.assertEqual(view["Content-Type"], "image/png")
            view.close()

            self.client.force_login(self.other)
            self.assertEqual(self.client.get(self.record.data["IMAGE_LINK"]).status_code, 404)
            self.assertEqual(
                self.client.post(url, {"image": self.image_upload("PNG", "green")}).status_code,
                404,
            )

    def test_rejects_non_image_upload(self):
        with tempfile.TemporaryDirectory() as temporary_root, override_settings(
            INVENTORY_UPLOAD_ROOT=Path(temporary_root)
        ):
            self.client.force_login(self.user)
            response = self.client.post(
                reverse("api_record_image", args=["sign", 1]),
                {"image": SimpleUploadedFile("fake.png", b"not an image", "image/png")},
            )
            self.assertEqual(response.status_code, 400)
            self.assertIn("valid image", response.json()["error"])

    def test_all_inventory_specs_and_form_template_expose_image_upload(self):
        self.assertEqual(settings.MAX_INVENTORY_IMAGE_BYTES, 20 * 1024 * 1024)
        self.assertGreater(settings.DATA_UPLOAD_MAX_MEMORY_SIZE, settings.MAX_INVENTORY_IMAGE_BYTES)
        self.client.force_login(self.user)
        for section in ("sign", "pavement", "lane", "curb"):
            response = self.client.get(reverse("api_spec", args=[section]))
            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertIn("IMAGE_LINK", payload["columns"])
            self.assertEqual(payload["columns"][-1], "IMAGE_LINK")
            self.assertEqual(payload["image_field"], "IMAGE_LINK")
        home = self.client.get(reverse("home"))
        self.assertContains(home, "Browse image")
        self.assertContains(home, "image-upload-input")
        self.assertContains(home, "MAX_IMAGE_UPLOAD_BYTES = 20 * 1024 * 1024")
        self.assertContains(home, "accessibleImageUrl")
        self.assertContains(home, "option.hidden = !option.dataset.projectName")
        self.assertContains(home, "updateControl.removeAttribute('title')")
        self.assertContains(home, "is-permission-disabled")
        self.assertIn(
            ".project-menu__option[hidden] { display:none!important; }",
            (settings.BASE_DIR / "static/inventory/css/brand.css").read_text(encoding="utf-8"),
        )
        self.assertContains(home, 'id="image-lightbox"')
        self.assertContains(home, "Image detail viewer")
        self.assertContains(home, "openImageLightbox")
        self.assertContains(home, "setLightboxScale")
        self.assertNotContains(home, 'id="swap-records-btn"')
        self.assertContains(home, "Insert below")
        self.assertContains(home, "beginInsertBelow")
        self.assertContains(home, "insert_after_id")
        self.assertIn("img-src 'self' data: blob:", home["Content-Security-Policy"])


class RecordOrderingTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("swap-user", password="StrongPass!234")
        self.other = User.objects.create_user("swap-other", password="StrongPass!234")
        self.admin = User.objects.create_superuser(
            "swap-admin", "swap-admin@example.com", "StrongPass!234"
        )
        self.project = Project.objects.create(name="Swap Project", code="swap-project")
        self.project.members.add(self.user, self.other)
        self.first = TabRecord.objects.create(
            project=self.project, owner=self.user, tab="sign", tab_record_id=1,
            data={"ST_ID": "ONE", "POLE_ID": "P1", "SIGN": "S1"},
        )
        self.second = TabRecord.objects.create(
            project=self.project, owner=self.user, tab="sign", tab_record_id=2,
            data={"ST_ID": "TWO", "POLE_ID": "P2", "SIGN": "S2"},
        )
        self.other_record = TabRecord.objects.create(
            project=self.project, owner=self.other, tab="sign", tab_record_id=3,
            data={"ST_ID": "THREE", "POLE_ID": "P3", "SIGN": "S3"},
        )

    def listed_ids(self):
        response = self.client.get(reverse("api_records", args=["sign"]))
        self.assertEqual(response.status_code, 200)
        return [record["ID"] for record in response.json()["records"]]

    def insert_below(self, target_id, st_id="INSERTED"):
        return self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({
                "row": {"ST_ID": st_id, "POLE_ID": "P4", "SIGN": "S4"},
                "insert_after_id": target_id,
            }),
            content_type="application/json",
        )

    def test_regular_user_inserts_below_own_record_without_changing_ids(self):
        self.client.force_login(self.user)
        response = self.insert_below(1)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["record"]["ID"], 4)
        self.assertEqual(self.listed_ids(), [1, 4, 2])
        self.assertEqual(
            list(
                TabRecord.objects.filter(project=self.project, tab="sign")
                .order_by("display_order")
                .values_list("tab_record_id", flat=True)
            ),
            [1, 4, 2, 3],
        )

    def test_regular_user_cannot_insert_below_another_users_record(self):
        self.client.force_login(self.user)
        response = self.insert_below(3)
        self.assertEqual(response.status_code, 404)
        self.assertEqual(self.listed_ids(), [1, 2])
        self.assertEqual(TabRecord.objects.count(), 3)

    def test_admin_can_insert_below_another_users_record(self):
        self.client.force_login(self.admin)
        response = self.insert_below(3, st_id="ADMIN-INSERT")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["record"]["ID"], 4)
        self.assertEqual(self.listed_ids(), [1, 2, 3, 4])


class AuthenticationTests(TestCase):
    def setUp(self):
        self.client = Client(enforce_csrf_checks=True)
        self.user = User.objects.create_user(
            username="inventoryuser",
            email="existing@example.com",
            password="StrongPass!234",
        )
        self.project = Project.objects.create(name="Authentication Project", code="auth-project")
        self.project.members.add(self.user)

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get(reverse("home"))
        self.assertRedirects(response, f"{reverse('login')}?next=/")

        response = self.client.get(reverse("api_spec", args=["sign"]))
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.json(), {"error": "Authentication required."})

    def test_login_grants_access_and_logout_revokes_it(self):
        response = self.client.get(reverse("login"))
        self.assertContains(response, "inventory/js/password-visibility")
        self.assertContains(response, "inventory/img/favicon-32x32")
        self.assertContains(response, "inventory/img/apple-touch-icon")
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("login"),
            {"username": "inventoryuser", "password": "StrongPass!234"},
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertRedirects(response, reverse("home"))
        home_response = self.client.get(reverse("home"))
        self.assertEqual(home_response.status_code, 200)
        self.assertContains(home_response, "MM-DD-YYYY")
        self.assertContains(home_response, "date-picker-btn")
        self.assertContains(home_response, "Choose date")
        self.assertContains(home_response, "saveCurrentDraft")
        self.assertContains(home_response, "restoreCurrentDraft")
        self.assertNotContains(home_response, "clearForm(true)")
        self.assertIn("default-src 'self'", home_response["Content-Security-Policy"])
        self.assertEqual(home_response["Cache-Control"], "no-store, private")
        self.assertIn("microphone=(self)", home_response["Permissions-Policy"])


    def test_session_timeout_defaults_are_security_focused(self):
        self.assertEqual(settings.SESSION_COOKIE_AGE, 8 * 60 * 60)
        self.assertTrue(settings.SESSION_EXPIRE_AT_BROWSER_CLOSE)
        self.assertTrue(settings.SESSION_SAVE_EVERY_REQUEST)

    @override_settings(SESSION_COOKIE_AGE=3600)
    def test_inactive_session_is_rejected(self):
        self.client.force_login(self.user)
        session = self.client.session
        session[LAST_ACTIVITY_KEY] = int(
            (timezone.now() - timedelta(seconds=3601)).timestamp()
        )
        session.save()

        response = self.client.get(reverse("home"))

        self.assertRedirects(response, f"{reverse('login')}?next=/")

    @override_settings(SESSION_COOKIE_AGE=3600)
    def test_legacy_two_week_session_is_rejected(self):
        self.client.force_login(self.user)
        session = self.client.session
        session.set_expiry(14 * 24 * 60 * 60)
        session.pop(LAST_ACTIVITY_KEY, None)
        session.save()

        response = self.client.get(reverse("home"))

        self.assertRedirects(response, f"{reverse('login')}?next=/")

    @override_settings(SESSION_COOKIE_AGE=3600)
    def test_recent_session_remains_authenticated(self):
        self.client.force_login(self.user)
        session = self.client.session
        session[LAST_ACTIVITY_KEY] = int(
            (timezone.now() - timedelta(seconds=60)).timestamp()
        )
        session.save()

        response = self.client.get(reverse("home"))

        self.assertEqual(response.status_code, 200)

    def test_assistant_preview_requires_authentication(self):
        csrf = self.client.get(reverse("login")).cookies["csrftoken"].value
        response = self.client.post(
            reverse("api_assistant_preview"),
            data='{"prompt":"add a sign"}',
            content_type="application/json",
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 401)

    @patch("inventory.views.interpret_inventory_prompt")
    def test_authenticated_user_can_preview_without_saving(self, interpret):
        interpret.return_value = {
            "section": "sign",
            "section_label": "Sign Inventory",
            "row": {"ST_ID": "100", "POLE_ID": "P2", "SIGN": "R1-1"},
            "summary": "Add one sign record.",
        }
        self.client.force_login(self.user)
        csrf = self.client.get(reverse("home")).cookies["csrftoken"].value
        response = self.client.post(
            reverse("api_assistant_preview"),
            data='{"prompt":"add sign 100 pole P2 R1-1"}',
            content_type="application/json",
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["section"], "sign")
        self.assertEqual(response.json()["row"]["SIGN_UID"], "SR_100_P2_R1-1")
        self.assertEqual(TabRecord.objects.count(), 0)

    def test_assistant_rejects_empty_prompt(self):
        self.client.force_login(self.user)
        csrf = self.client.get(reverse("home")).cookies["csrftoken"].value
        response = self.client.post(
            reverse("api_assistant_preview"),
            data='{"prompt":"  "}',
            content_type="application/json",
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 400)

    @patch("inventory.views.transcribe_audio", return_value="Add sign record ST ID 100")
    def test_authenticated_user_can_transcribe_voice(self, transcribe):
        self.client.force_login(self.user)
        csrf = self.client.get(reverse("home")).cookies["csrftoken"].value
        audio = SimpleUploadedFile(
            "voice.webm",
            b"test-audio",
            content_type="audio/webm",
        )
        response = self.client.post(
            reverse("api_assistant_transcribe"),
            {"audio": audio},
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["text"], "Add sign record ST ID 100")
        transcribe.assert_called_once()

        csrf = self.client.cookies["csrftoken"].value
        response = self.client.post(
            reverse("logout"), HTTP_X_CSRFTOKEN=csrf
        )
        self.assertRedirects(response, reverse("login"))
        self.assertEqual(self.client.get(reverse("home")).status_code, 302)

    def test_voice_transcription_rejects_non_audio_uploads(self):
        self.client.force_login(self.user)
        csrf = self.client.get(reverse("home")).cookies["csrftoken"].value
        upload = SimpleUploadedFile(
            "instructions.txt",
            b"not an audio recording",
            content_type="text/plain",
        )
        response = self.client.post(
            reverse("api_assistant_transcribe"),
            {"audio": upload},
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("supported voice recording", response.json()["error"])

    def test_signup_creates_inactive_user_pending_admin_approval(self):
        response = self.client.get(reverse("signup"))
        self.assertContains(response, "inventory/js/password-visibility")
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("signup"),
            {
                "username": "newuser",
                "email": "new@example.com",
                "password1": "AnotherStrong!234",
                "password2": "AnotherStrong!234",
            },
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertRedirects(response, reverse("signup_pending"))
        user = User.objects.get(username="newuser", email="new@example.com")
        self.assertFalse(user.is_active)
        self.assertContains(self.client.get(reverse("signup_pending")), "pending administrator approval")
        self.assertContains(self.client.get(reverse("signup_pending")), "auth-status-card")
        self.assertContains(self.client.get(reverse("signup_pending")), "Return to login")
        self.assertRedirects(self.client.get(reverse("home")), f"{reverse('login')}?next=/")

    def test_inactive_user_gets_pending_or_deactivated_login_message(self):
        self.user.is_active = False
        self.user.save(update_fields=["is_active"])
        response = self.client.get(reverse("login"))
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("login"),
            {"username": self.user.username, "password": "StrongPass!234"},
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "pending administrator approval or has been deactivated")
        self.assertNotIn("_auth_user_id", self.client.session)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        APP_BASE_URL="https://inventory.example.com",
    )
    def test_admin_approval_sends_email_and_enables_login(self):
        self.user.is_active = False
        self.user.save(update_fields=["is_active"])

        with self.captureOnCommitCallbacks(execute=True):
            self.user.is_active = True
            self.user.save(update_fields=["is_active"])

        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].subject, "Bluedome Inventory account approved")
        self.assertIn("https://inventory.example.com/login/", mail.outbox[0].body)
        self.assertTrue(self.client.login(username=self.user.username, password="StrongPass!234"))

        self.user.is_active = False
        self.user.save(update_fields=["is_active"])
        self.assertRedirects(self.client.get(reverse("home")), f"{reverse('login')}?next=/")
        self.client.logout()
        self.assertFalse(self.client.login(username=self.user.username, password="StrongPass!234"))

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_admin_user_list_has_visible_activate_and_deactivate_buttons(self):
        admin_user = User.objects.create_superuser(
            username="approvaladmin",
            email="approvaladmin@example.com",
            password="StrongPass!234",
        )
        pending = User.objects.create_user(
            username="pendinguser",
            email="pending@example.com",
            password="StrongPass!234",
            is_active=False,
        )
        self.client.force_login(admin_user)

        changelist = self.client.get(reverse("admin:auth_user_changelist"))
        self.assertContains(changelist, "Activate")
        self.assertContains(changelist, "Deactivate")
        self.assertContains(changelist, "Pending / inactive")

        activate_url = reverse("admin:auth_user_activate_account", args=[pending.pk])
        confirmation = self.client.get(activate_url)
        self.assertContains(confirmation, "Activate user")
        self.assertContains(confirmation, "account-confirmation-summary")
        self.assertContains(confirmation, "Current status")
        csrf = self.client.cookies["csrftoken"].value
        with self.captureOnCommitCallbacks(execute=True):
            response = self.client.post(activate_url, HTTP_X_CSRFTOKEN=csrf)
        self.assertRedirects(response, reverse("admin:auth_user_changelist"))
        pending.refresh_from_db()
        self.assertTrue(pending.is_active)
        self.assertEqual(len(mail.outbox), 1)

        deactivate_url = reverse("admin:auth_user_deactivate_account", args=[self.user.pk])
        confirmation = self.client.get(deactivate_url)
        self.assertContains(confirmation, "Deactivate user")
        csrf = self.client.cookies["csrftoken"].value
        response = self.client.post(deactivate_url, HTTP_X_CSRFTOKEN=csrf)
        self.assertRedirects(response, reverse("admin:auth_user_changelist"))
        self.user.refresh_from_db()
        self.assertFalse(self.user.is_active)
        self.assertEqual(len(mail.outbox), 2)
        self.assertEqual(mail.outbox[1].subject, "Bluedome Inventory account deactivated")
        self.assertIn("can no longer access", mail.outbox[1].body)

    def test_admin_user_change_hides_password_hash_metadata(self):
        admin_user = User.objects.create_superuser(
            username="securityadmin",
            email="securityadmin@example.com",
            password="StrongPass!234",
        )
        self.client.force_login(admin_user)

        response = self.client.get(
            reverse("admin:auth_user_change", args=[self.user.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Password is securely stored")
        self.assertContains(response, "Reset password")
        self.assertNotContains(response, "algorithm:")
        self.assertNotContains(response, "iterations:")

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.console.EmailBackend")
    def test_admin_activation_warns_when_smtp_is_not_configured(self):
        admin_user = User.objects.create_superuser(
            username="emailadmin",
            email="emailadmin@example.com",
            password="StrongPass!234",
        )
        pending = User.objects.create_user(
            username="emailpending",
            email="emailpending@example.com",
            password="StrongPass!234",
            is_active=False,
        )
        self.client.force_login(admin_user)
        activate_url = reverse("admin:auth_user_activate_account", args=[pending.pk])
        self.client.get(activate_url)
        csrf = self.client.cookies["csrftoken"].value
        response = self.client.post(
            activate_url,
            HTTP_X_CSRFTOKEN=csrf,
            follow=True,
        )
        self.assertContains(response, "SMTP is not configured")
        pending.refresh_from_db()
        self.assertTrue(pending.is_active)

    def test_duplicate_email_is_rejected(self):
        response = self.client.get(reverse("signup"))
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("signup"),
            {
                "username": "anotheruser",
                "email": "EXISTING@example.com",
                "password1": "AnotherStrong!234",
                "password2": "AnotherStrong!234",
            },
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "An account with this email already exists.")
        self.assertFalse(User.objects.filter(username="anotheruser").exists())

    def test_signup_rejects_email_without_valid_domain_pattern(self):
        response = self.client.get(reverse("signup"))
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("signup"),
            {
                "username": "invalidemailuser",
                "email": "person@localhost",
                "password1": "AnotherStrong!234",
                "password2": "AnotherStrong!234",
            },
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Enter a valid email address")
        self.assertFalse(User.objects.filter(username="invalidemailuser").exists())

    def test_password_requirement_endpoint_uses_django_validators(self):
        response = self.client.get(reverse("signup"))
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("password_requirements"),
            data=(
                '{"username":"newuser","email":"new@example.com",'
                '"password":"newuser123"}'
            ),
            content_type="application/json",
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 200)
        checks = response.json()["checks"]
        self.assertFalse(checks["UserAttributeSimilarityValidator"])
        self.assertTrue(checks["MinimumLengthValidator"])
        self.assertTrue(checks["NumericPasswordValidator"])

        response = self.client.post(
            reverse("password_requirements"),
            data='{"username":"newuser","email":"new@example.com","password":"12345678"}',
            content_type="application/json",
            HTTP_X_CSRFTOKEN=csrf,
        )
        checks = response.json()["checks"]
        self.assertFalse(checks["NumericPasswordValidator"])
        self.assertFalse(checks["CommonPasswordValidator"])

    def test_write_api_requires_csrf_token(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("api_records", args=["sign"]),
            data='{"row": {}}',
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 403)

    def test_endpoints_reject_unsupported_methods(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("home"))
        csrf = response.cookies["csrftoken"].value
        self.assertEqual(
            self.client.post(reverse("home"), HTTP_X_CSRFTOKEN=csrf).status_code,
            405,
        )
        self.assertEqual(
            self.client.post(
                reverse("api_spec", args=["sign"]),
                HTTP_X_CSRFTOKEN=csrf,
            ).status_code,
            405,
        )
        self.assertEqual(
            self.client.post(
                reverse("api_export_all"),
                HTTP_X_CSRFTOKEN=csrf,
            ).status_code,
            405,
        )

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_password_reset_changes_password_with_single_use_token(self):
        response = self.client.get(reverse("password_reset"))
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("password_reset"),
            {"email": "existing@example.com"},
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertRedirects(response, reverse("password_reset_done"))
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].subject, "Bluedome Inventory password reset")
        self.assertNotIn("StrongPass!234", mail.outbox[0].body)

        match = re.search(r"http://testserver(\S+)", mail.outbox[0].body)
        self.assertIsNotNone(match)
        token_url = match.group(1)

        response = self.client.get(token_url)
        self.assertEqual(response.status_code, 302)
        set_password_url = response.url
        response = self.client.get(set_password_url)
        self.assertEqual(response.status_code, 200)

        csrf = self.client.cookies["csrftoken"].value
        response = self.client.post(
            set_password_url,
            {
                "new_password1": "ReplacementPass!456",
                "new_password2": "ReplacementPass!456",
            },
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertRedirects(response, reverse("password_reset_complete"))

        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password("ReplacementPass!456"))
        self.assertFalse(self.user.check_password("StrongPass!234"))

        response = self.client.get(token_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Link expired")

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_password_reset_does_not_reveal_unknown_email(self):
        response = self.client.get(reverse("password_reset"))
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("password_reset"),
            {"email": "unknown@example.com"},
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertRedirects(response, reverse("password_reset_done"))
        self.assertEqual(len(mail.outbox), 0)


class DatabaseConfigurationTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username="regularuser",
            password="StrongPass!234",
        )
        self.admin_user = User.objects.create_superuser(
            username="sectionadmin",
            email="admin@example.com",
            password="StrongPass!234",
        )
        self.project = Project.objects.create(name="Test Project", code="test-project")
        self.project.members.add(self.user)

    def test_runtime_options_come_from_database(self):
        section = InventorySection.objects.get(key="sign")
        DropdownOption.objects.create(
            section=section,
            field_name="SIGN_CONDITION",
            value="ADMIN ADDED VALUE",
            sort_order=999,
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("api_spec", args=["sign"]))
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "ADMIN ADDED VALUE",
            response.json()["options"]["SIGN_CONDITION"],
        )

    def test_regular_users_only_see_and_modify_their_own_project_records(self):
        other_user = User.objects.create_user(
            username="otheruser",
            password="StrongPass!234",
        )
        own = TabRecord.objects.create(
            project=self.project,
            owner=self.user,
            tab="sign",
            tab_record_id=1,
            data={"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1"},
        )
        other = TabRecord.objects.create(
            project=self.project,
            owner=other_user,
            tab="sign",
            tab_record_id=2,
            data={"ST_ID": "200", "POLE_ID": "P2", "SIGN": "S2"},
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("api_records", args=["sign"]))
        self.assertEqual(
            [row["ID"] for row in response.json()["records"]],
            [own.tab_record_id],
        )
        records = response.json()["records"]
        self.assertTrue(records[0]["_IS_OWN"])
        self.assertTrue(records[0]["_CAN_EDIT"])
        self.assertEqual(records[0]["_OWNER_NAME"], "regularuser")

        response = self.client.delete(
            reverse("api_record_detail", args=["sign", other.tab_record_id])
        )
        self.assertEqual(response.status_code, 404)
        self.assertTrue(TabRecord.objects.filter(pk=other.pk).exists())

        response = self.client.put(
            reverse("api_record_detail", args=["sign", other.tab_record_id]),
            data=json.dumps({"row": {"ST_ID": "CHANGED", "POLE_ID": "P2", "SIGN": "S2"}}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 404)
        other.refresh_from_db()
        self.assertEqual(other.data["ST_ID"], "200")

        response = self.client.put(
            reverse("api_record_detail", args=["sign", own.tab_record_id]),
            data=json.dumps({"row": {"ST_ID": "OWN-EDIT", "POLE_ID": "P1", "SIGN": "S1"}}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        own.refresh_from_db()
        self.assertEqual(own.data["ST_ID"], "OWN-EDIT")

        response = self.client.delete(reverse("api_records", args=["sign"]))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(TabRecord.objects.filter(pk=own.pk).exists())
        self.assertTrue(TabRecord.objects.filter(pk=other.pk).exists())

    def test_admin_sees_all_inventory_with_owner_labels(self):
        other_user = User.objects.create_user(
            username="fieldworker",
            first_name="Field",
            last_name="Worker",
            password="StrongPass!234",
        )
        TabRecord.objects.create(
            project=self.project,
            owner=self.user,
            tab="sign",
            tab_record_id=1,
            data={"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1"},
        )
        TabRecord.objects.create(
            project=self.project,
            owner=other_user,
            tab="sign",
            tab_record_id=2,
            data={"ST_ID": "200", "POLE_ID": "P2", "SIGN": "S2"},
        )
        self.client.force_login(self.admin_user)
        spec_response = self.client.get(reverse("api_spec", args=["sign"]))
        self.assertNotIn("ADDED_BY", spec_response.json()["columns"])
        response = self.client.get(reverse("api_records", args=["sign"]))
        records = response.json()["records"]
        self.assertEqual(len(records), 2)
        self.assertNotIn("ADDED_BY", records[0])
        self.assertNotIn("ADDED_BY", records[1])
        self.assertTrue(all(record["_CAN_EDIT"] for record in records))
        self.assertEqual(
            [record["_OWNER_NAME"] for record in records],
            ["regularuser", "Field Worker"],
        )

        admin_response = self.client.get(
            reverse("admin:inventory_tabrecord_changelist")
        )
        self.assertContains(admin_response, "Username")
        self.assertContains(admin_response, "Field Worker")
        self.assertContains(admin_response, "Date")

        dashboard_response = self.client.get(reverse("admin:index"))
        self.assertContains(dashboard_response, "Inventory Records")
        self.assertContains(dashboard_response, "Field Worker")
        self.assertContains(dashboard_response, ">1</a>", html=False)
        self.assertContains(dashboard_response, "Section")
        self.assertContains(dashboard_response, "Date added")
        self.assertContains(
            dashboard_response,
            TabRecord.objects.order_by("pk").first().created_at.strftime("%m-%d-%Y"),
        )

        self.assertContains(admin_response, "By Username")
        self.assertContains(admin_response, "By Inventory section")
        username_response = self.client.get(
            reverse("admin:inventory_tabrecord_changelist"),
            {"username": other_user.pk},
        )
        self.assertContains(username_response, "Field Worker")
        self.assertEqual(username_response.context["cl"].result_count, 1)
        section_response = self.client.get(
            reverse("admin:inventory_tabrecord_changelist"),
            {"inventory_section": "sign"},
        )
        self.assertEqual(section_response.context["cl"].result_count, 2)
        section_filter = next(
            spec
            for spec in section_response.context["cl"].filter_specs
            if spec.title == "Inventory section"
        )
        sign_choices = [
            choice for choice in section_filter.lookup_choices if choice[0] == "sign"
        ]
        self.assertEqual(len(sign_choices), 1)

    def test_admin_can_update_and_delete_another_users_record(self):
        record = TabRecord.objects.create(
            project=self.project,
            owner=self.user,
            tab="sign",
            tab_record_id=1,
            data={"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1"},
        )
        self.client.force_login(self.admin_user)

        updated = self.client.put(
            reverse("api_record_detail", args=["sign", record.tab_record_id]),
            data=json.dumps({"row": {"ST_ID": "ADMIN-EDIT", "POLE_ID": "P1", "SIGN": "S1"}}),
            content_type="application/json",
        )
        self.assertEqual(updated.status_code, 200)
        record.refresh_from_db()
        self.assertEqual(record.data["ST_ID"], "ADMIN-EDIT")

        deleted = self.client.delete(
            reverse("api_record_detail", args=["sign", record.tab_record_id])
        )
        self.assertEqual(deleted.status_code, 200)
        self.assertFalse(TabRecord.objects.filter(pk=record.pk).exists())

    def test_new_inventory_record_is_assigned_to_authenticated_user(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({
                "row": {"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1"}
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(TabRecord.objects.get().owner, self.user)

    def test_admin_can_filter_inventory_by_date_added_range(self):
        older = TabRecord.objects.create(
            project=self.project,
            owner=self.user,
            tab="sign",
            tab_record_id=1,
            data={"ST_ID": "OLD", "POLE_ID": "P1", "SIGN": "S1"},
        )
        newer = TabRecord.objects.create(
            project=self.project,
            owner=self.user,
            tab="sign",
            tab_record_id=2,
            data={"ST_ID": "NEW", "POLE_ID": "P2", "SIGN": "S2"},
        )
        target_date = timezone.localdate() - timedelta(days=2)
        older_date = target_date - timedelta(days=3)
        TabRecord.objects.filter(pk=older.pk).update(
            created_at=timezone.make_aware(
                datetime.combine(older_date, datetime.min.time())
            )
        )
        TabRecord.objects.filter(pk=newer.pk).update(
            created_at=timezone.make_aware(
                datetime.combine(target_date, datetime.min.time())
            )
        )

        self.client.force_login(self.admin_user)
        response = self.client.get(
            reverse("admin:inventory_tabrecord_changelist"),
            {
                "date_from": target_date.strftime("%m-%d-%Y"),
                "date_to": target_date.strftime("%m-%d-%Y"),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Date added")
        self.assertContains(response, "MM-DD-YYYY")
        self.assertContains(response, 'type="date"')
        self.assertContains(response, "Choose From date")
        self.assertContains(response, "Choose To date")
        self.assertEqual(response.context["cl"].result_count, 1)
        self.assertEqual(response.context["cl"].result_list[0].pk, newer.pk)

    def test_inventory_dates_require_mm_dd_yyyy(self):
        self.client.force_login(self.user)
        row = {
            "ST_ID": "100",
            "POLE_ID": "P1",
            "SIGN": "S1",
            "INSP_DATE": "21-07-2026",
        }
        response = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({"row": row}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("MM-DD-YYYY", response.json()["error"])

        row["INSP_DATE"] = "07-21-2026"
        response = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({"row": row}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(TabRecord.objects.get().data["INSP_DATE"], "07-21-2026")

    def test_coordinates_are_rounded_to_seven_decimal_places(self):
        self.client.force_login(self.user)
        row = {
            "ST_ID": "100",
            "POLE_ID": "P1",
            "SIGN": "S1",
            "LATITUDE": "12.34567894",
            "LONGITUDE": "-98.76543216",
        }
        response = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({"row": row}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        stored = TabRecord.objects.get().data
        self.assertEqual(stored["LATITUDE"], "12.3456789")
        self.assertEqual(stored["LONGITUDE"], "-98.7654322")

        row["LATITUDE"] = "91.0000000"
        response = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({"row": row}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("between -90 and 90", response.json()["error"])

    def test_admin_mutcd_mapping_populates_both_database_dropdowns(self):
        section = InventorySection.objects.get(key="sign")
        MutcdMapping.objects.create(
            section=section,
            word_description="testdesc",
            mutcd_code="testcode",
            classification="testcl",
        )
        MutcdMapping.objects.create(
            section=section,
            word_description="testdesc alternate",
            mutcd_code="testcode",
            classification="alternate class",
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("api_spec", args=["sign"]))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("testcode", data["options"]["MUTCD"])
        self.assertIn("testdesc", data["options"]["WORD_DESCRIPTION"])
        self.assertEqual(data["mutcd_to_class"]["testcode"], "testcl")
        self.assertEqual(data["mutcd_map"]["testdesc"]["MUTCD"], "testcode")
        self.assertEqual(
            data["mutcd_word_options"]["testcode"],
            ["testdesc", "testdesc alternate"],
        )
        self.assertIn("no-store", response["Cache-Control"])

        save_response = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({
                "row": {
                    "ST_ID": "100",
                    "POLE_ID": "P1",
                    "SIGN": "S1",
                    "MUTCD": "testcode",
                    "WORD_DESCRIPTION": "testdesc alternate",
                }
            }),
            content_type="application/json",
        )
        self.assertEqual(save_response.status_code, 200)
        self.assertEqual(
            save_response.json()["record"]["MUTCD_CLASSIFICATION"],
            "alternate class",
        )

    def test_all_database_dropdown_values_are_alphabetically_sorted(self):
        self.client.force_login(self.user)
        for section_key in ("sign", "pavement", "lane", "curb"):
            response = self.client.get(reverse("api_spec", args=[section_key]))
            self.assertEqual(response.status_code, 200)
            for field_name, values in response.json()["options"].items():
                self.assertEqual(
                    values,
                    sorted(values, key=lambda value: str(value).casefold()),
                    msg=f"{section_key}.{field_name} is not alphabetically sorted",
                )

    def test_only_admin_role_can_open_configuration_admin(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("admin:index"))
        self.assertEqual(response.status_code, 302)

        self.client.force_login(self.admin_user)
        self.assertEqual(self.client.get(reverse("admin:index")).status_code, 200)
        self.assertEqual(
            self.client.get(
                reverse("admin:inventory_dropdownoption_changelist")
            ).status_code,
            200,
        )

    def test_admin_pages_use_bluedome_branding(self):
        login_response = self.client.get(reverse("admin:login"))
        self.assertContains(login_response, "Bluedome Inventory")
        self.assertNotContains(login_response, "Django administration")

        self.client.force_login(self.admin_user)
        index_response = self.client.get(reverse("admin:index"))
        self.assertContains(index_response, "Bluedome Inventory")
        self.assertContains(index_response, "Inventory Configuration")
        self.assertNotContains(index_response, "Django administration")

    def test_configuration_rows_have_edit_and_delete_actions(self):
        self.client.force_login(self.admin_user)
        option = DropdownOption.objects.create(
            section_id="sign",
            field_name="SIGN_CONDITION",
            value="ROW ACTION TEST VALUE",
            sort_order=9999,
        )
        response = self.client.get(
            reverse("admin:inventory_dropdownoption_changelist"),
            {"q": "ROW ACTION TEST VALUE"},
        )
        self.assertContains(
            response,
            reverse(
                "admin:inventory_dropdownoption_change",
                args=[option.pk],
            ),
        )
        self.assertContains(
            response,
            reverse(
                "admin:inventory_dropdownoption_delete",
                args=[option.pk],
            ),
        )
        self.assertContains(response, "Edit")
        self.assertContains(response, "Delete")
        self.assertContains(response, 'aria-label="Edit"')
        self.assertContains(response, 'aria-label="Delete"')

        change_response = self.client.get(
            reverse("admin:inventory_dropdownoption_change", args=[option.pk])
        )
        self.assertContains(change_response, "icon-delete")
        self.assertContains(change_response, "<svg", html=False)

    def test_admin_can_delete_mutcd_mapping_with_confirmation(self):
        mapping = MutcdMapping.objects.create(
            section_id="sign",
            word_description="DELETE TEST DESCRIPTION",
            mutcd_code="DELETE-TEST-CODE",
            classification="TEST",
        )
        self.client.force_login(self.admin_user)
        delete_url = reverse(
            "admin:inventory_mutcdmapping_delete",
            args=[mapping.pk],
        )
        confirmation = self.client.get(delete_url)
        self.assertEqual(confirmation.status_code, 200)
        self.assertContains(confirmation, "Delete mutcd mapping?")
        self.assertContains(confirmation, "Delete permanently")
        self.assertContains(confirmation, "button-danger")
        self.assertContains(confirmation, "button-secondary")

        response = self.client.post(delete_url, {"post": "yes"})
        self.assertRedirects(
            response,
            reverse("admin:inventory_mutcdmapping_changelist"),
        )
        self.assertFalse(MutcdMapping.objects.filter(pk=mapping.pk).exists())

    @staticmethod
    def _excel_upload(section_key, rows):
        from .specs import get_spec

        spec = get_spec(section_key)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(spec["columns"])
        for values in rows:
            worksheet.append([values.get(column, "") for column in spec["columns"]])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return SimpleUploadedFile(
            "inventory-import.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_admin_can_bulk_import_valid_excel_records(self):
        self.client.force_login(self.admin_user)
        upload = self._excel_upload(
            "sign",
            [
                {"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1", "STREET_NAME": "FIRST ST"},
                {"ST_ID": "100", "POLE_ID": "P2", "SIGN": "S2", "STREET_NAME": "SECOND ST"},
            ],
        )
        response = self.client.post(
            reverse("admin:inventory_tabrecord_import_excel"),
            {"section": "sign", "excel_file": upload},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Successfully imported 2 Sign Inventory records")
        imported = list(TabRecord.objects.filter(tab="sign").order_by("tab_record_id"))
        self.assertEqual(len(imported), 2)
        self.assertEqual(imported[0].data["SIGN_UID"], "SR_100_P1_S1")
        self.assertEqual(imported[1].data["SIGN_UID"], "SR_100_P2_S2")

    def test_admin_can_bulk_import_mutcd_mappings(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Section", "Word description", "MUTCD code", "Classification"])
        worksheet.append(["sign", "Bulk sign one", "B-1", "Warning"])
        worksheet.append(["sign", "Bulk sign two", "B-2", "Regulatory"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "mutcd-mappings.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_login(self.admin_user)
        response = self.client.post(
            reverse("admin:inventory_mutcdmapping_import_excel"),
            {"excel_file": upload},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Successfully imported 2 mutcd mappings")
        self.assertTrue(
            MutcdMapping.objects.filter(
                section_id="sign",
                word_description="Bulk sign one",
                mutcd_code="B-1",
            ).exists()
        )

    def test_mutcd_excel_import_updates_existing_and_adds_new_without_duplicates(self):
        existing = MutcdMapping.objects.create(
            section_id="sign",
            word_description="Existing sign",
            mutcd_code="OLD",
            classification="Old classification",
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Section", "Word description", "MUTCD code", "Classification"])
        worksheet.append(["sign", "Existing sign", "UPDATED", "Updated classification"])
        worksheet.append(["sign", "New sign one", "NEW-1", "New classification"])
        worksheet.append(["sign", "New sign two", "NEW-2", "New classification"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "mutcd-upsert.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.force_login(self.admin_user)
        response = self.client.post(
            reverse("admin:inventory_mutcdmapping_import_excel"),
            {"excel_file": upload},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Successfully imported 3 mutcd mappings")
        self.assertContains(response, "2 created, 1 updated")
        existing.refresh_from_db()
        self.assertEqual(existing.mutcd_code, "UPDATED")
        self.assertEqual(existing.classification, "Updated classification")
        self.assertEqual(
            MutcdMapping.objects.filter(
                section_id="sign",
                word_description="Existing sign",
            ).count(),
            1,
        )
        self.assertTrue(
            MutcdMapping.objects.filter(
                section_id="sign",
                word_description="New sign one",
                mutcd_code="NEW-1",
                classification="New classification",
            ).exists()
        )
        self.assertTrue(
            MutcdMapping.objects.filter(
                section_id="sign",
                word_description="New sign two",
                mutcd_code="NEW-2",
                classification="New classification",
            ).exists()
        )

    def test_mutcd_import_syncs_classification_and_deterministic_fallback(self):
        MutcdFallback.objects.create(
            section_id="sign",
            code="W13-1P-TEST",
            word_description="Old fallback",
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Section", "Word description", "MUTCD code", "Classification"])
        worksheet.append(["sign", "10 M P H TEST", "W13-1P-TEST", "WARNING SIGN"])
        worksheet.append(["sign", "25 M P H TEST", "W13-1P-TEST", "WARNING SIGN"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "mutcd-related-sync.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.client.force_login(self.admin_user)
        response = self.client.post(
            reverse("admin:inventory_mutcdmapping_import_excel"),
            {"excel_file": upload},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            MutcdFallback.objects.get(
                section_id="sign", code="W13-1P-TEST"
            ).word_description,
            "10 M P H TEST",
        )
        self.assertEqual(
            MutcdClassification.objects.get(
                section_id="sign", code="W13-1P-TEST"
            ).classification,
            "WARNING SIGN",
        )
        self.assertEqual(
            MutcdMapping.objects.filter(
                section_id="sign", mutcd_code="W13-1P-TEST"
            ).count(),
            2,
        )
    def test_empty_configuration_workbook_is_rejected(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Section", "Word description", "MUTCD code", "Classification"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "empty-mutcd-mappings.xlsx",
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.client.force_login(self.admin_user)
        response = self.client.post(
            reverse("admin:inventory_mutcdmapping_import_excel"),
            {"excel_file": upload},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "The workbook does not contain any data rows.")

    def test_mutcd_mapping_admin_shows_bulk_excel_controls(self):
        self.client.force_login(self.admin_user)
        response = self.client.get(
            reverse("admin:inventory_mutcdmapping_changelist")
        )
        self.assertContains(response, "Import Excel")
        self.assertContains(response, "Download Template")
        self.assertContains(response, "Export All")
        add_response = self.client.get(
            reverse("admin:inventory_mutcdmapping_add")
        )
        self.assertContains(add_response, "Import Multiple via Excel")
        import_response = self.client.get(
            reverse("admin:inventory_mutcdmapping_import_excel")
        )
        self.assertContains(import_response, "Export All Data")
        self.assertContains(
            import_response,
            reverse("admin:inventory_mutcdmapping_export_excel"),
        )

    def test_invalid_excel_import_is_atomic(self):
        self.client.force_login(self.admin_user)
        upload = self._excel_upload(
            "sign",
            [
                {"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1"},
                {"ST_ID": "100", "POLE_ID": "", "SIGN": "S2"},
            ],
        )
        response = self.client.post(
            reverse("admin:inventory_tabrecord_import_excel"),
            {"section": "sign", "excel_file": upload},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Row 3: missing required values POLE_ID")
        self.assertEqual(TabRecord.objects.filter(tab="sign").count(), 0)

    @override_settings(MAX_XLSX_UNCOMPRESSED_BYTES=1)
    def test_excel_import_rejects_workbook_expansion_over_limit(self):
        self.client.force_login(self.admin_user)
        upload = self._excel_upload(
            "sign",
            [{"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1"}],
        )
        response = self.client.post(
            reverse("admin:inventory_tabrecord_import_excel"),
            {"section": "sign", "excel_file": upload},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "expands beyond the safe processing limit")
        self.assertFalse(TabRecord.objects.exists())

    def test_admin_excel_template_and_full_export(self):
        self.client.force_login(self.admin_user)
        template_response = self.client.get(
            reverse("admin:inventory_tabrecord_excel_template")
        )
        self.assertEqual(template_response.status_code, 200)
        template_book = load_workbook(io.BytesIO(template_response.content), read_only=True)
        self.assertEqual(
            template_book.sheetnames,
            ["Sign Inventory", "Pavement Inventory", "Lane Inventory", "Curb Inventory"],
        )
        self.assertEqual(template_book["Sign Inventory"]["A1"].value, "ID")

        TabRecord.objects.create(
            project=self.project,
            tab="sign",
            tab_record_id=1,
            data={"ST_ID": "100", "POLE_ID": "P1", "SIGN": "S1", "SIGN_UID": "SR_100_P1_S1"},
        )
        export_response = self.client.get(
            reverse("admin:inventory_tabrecord_export_excel")
        )
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_response["X-Record-Count"], "1")
        export_book = load_workbook(io.BytesIO(export_response.content), read_only=True)
        sign_sheet = export_book["Sign Inventory"]
        self.assertEqual(sign_sheet["A2"].value, 1)
        self.assertEqual(sign_sheet["B2"].value, "100")

    def test_excel_export_neutralizes_formula_cells(self):
        self.client.force_login(self.user)
        TabRecord.objects.create(
            project=self.project,
            tab="sign",
            tab_record_id=1,
            owner=self.user,
            data={
                "ST_ID": "100",
                "STREET_NAME": '=HYPERLINK("https://example.invalid","click")',
                "POLE_ID": "P1",
                "SIGN": "S1",
                "SIGN_UID": "SR_100_P1_S1",
            },
        )
        response = self.client.get(reverse("api_export", args=["sign"]))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content), data_only=False)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        street_name_column = headers.index("STREET_NAME") + 1
        self.assertTrue(sheet.cell(2, street_name_column).value.startswith("'="))

    def test_regular_user_cannot_access_admin_excel_tools(self):
        self.client.force_login(self.user)
        for url_name in (
            "admin:inventory_tabrecord_import_excel",
            "admin:inventory_tabrecord_excel_template",
            "admin:inventory_tabrecord_export_excel",
        ):
            response = self.client.get(reverse(url_name))
            self.assertEqual(response.status_code, 302)

    def test_only_admin_role_can_add_reference_options(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("api_options", args=["sign"]),
            data='{"field":"SIGN_CONDITION","value":"UNAUTHORIZED VALUE"}',
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(
            DropdownOption.objects.filter(value="UNAUTHORIZED VALUE").exists()
        )

        self.client.force_login(self.admin_user)
        response = self.client.post(
            reverse("api_options", args=["sign"]),
            data='{"field":"SIGN_CONDITION","value":"AUTHORIZED VALUE"}',
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            DropdownOption.objects.filter(value="AUTHORIZED VALUE").exists()
        )

        response = self.client.post(
            reverse("api_options", args=["sign"]),
            data='{"field":"LATITUDE","value":"SHOULD NOT BE CREATED"}',
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(
            DropdownOption.objects.filter(value="SHOULD NOT BE CREATED").exists()
        )

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_password_reset_rejects_invalid_email_pattern(self):
        response = self.client.get(reverse("password_reset"))
        csrf = response.cookies["csrftoken"].value
        response = self.client.post(
            reverse("password_reset"),
            {"email": "person@localhost"},
            HTTP_X_CSRFTOKEN=csrf,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Enter a valid email address")
        self.assertEqual(len(mail.outbox), 0)


class ProjectAccessTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("project-user", password="StrongPass!234")
        self.other = User.objects.create_user("other-project-user", password="StrongPass!234")
        self.admin = User.objects.create_superuser("project-admin", "admin@example.com", "StrongPass!234")
        self.alpha = Project.objects.create(name="Alpha", code="alpha")
        self.beta = Project.objects.create(name="Beta", code="beta")
        self.alpha.members.add(self.user)
        self.beta.members.add(self.other)
        TabRecord.objects.create(
            project=self.alpha, owner=self.user, tab="sign", tab_record_id=1,
            data={"ST_ID": "A", "POLE_ID": "P1", "SIGN": "S1"},
        )
        TabRecord.objects.create(
            project=self.beta, owner=self.other, tab="sign", tab_record_id=2,
            data={"ST_ID": "B", "POLE_ID": "P2", "SIGN": "S2"},
        )

    def test_user_only_sees_assigned_project_and_cannot_switch_to_another(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("api_records", args=["sign"]))
        self.assertEqual([row["ST_ID"] for row in response.json()["records"]], ["A"])
        forbidden = self.client.post(
            reverse("api_select_project"),
            data=json.dumps({"project_id": self.beta.pk}),
            content_type="application/json",
        )
        self.assertEqual(forbidden.status_code, 403)

    def test_admin_switches_project_context_without_combining_records(self):
        self.client.force_login(self.admin)
        self.client.post(
            reverse("api_select_project"),
            data=json.dumps({"project_id": self.beta.pk}),
            content_type="application/json",
        )
        response = self.client.get(reverse("api_records", args=["sign"]))
        self.assertEqual([row["ST_ID"] for row in response.json()["records"]], ["B"])

    def test_user_created_record_is_visible_to_admin_in_same_project(self):
        self.client.force_login(self.user)
        created = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({"row": {"ST_ID": "USER-NEW", "POLE_ID": "P9", "SIGN": "S9"}}),
            content_type="application/json",
        )
        self.assertEqual(created.status_code, 200)
        record = TabRecord.objects.get(project=self.alpha, data__ST_ID="USER-NEW")
        self.assertEqual(record.owner, self.user)

        self.client.force_login(self.admin)
        selected = self.client.post(
            reverse("api_select_project"),
            data=json.dumps({"project_id": self.alpha.pk}),
            content_type="application/json",
        )
        self.assertEqual(selected.status_code, 200)
        response = self.client.get(reverse("api_records", args=["sign"]))
        self.assertIn("USER-NEW", [row["ST_ID"] for row in response.json()["records"]])

    def test_opening_project_in_admin_syncs_view_site_project_context(self):
        self.client.force_login(self.admin)
        session = self.client.session
        session["inventory_project_id"] = self.beta.pk
        session.save()

        response = self.client.get(
            reverse("admin:inventory_project_change", args=[self.alpha.pk])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.client.session["inventory_project_id"], self.alpha.pk)
        view_site_records = self.client.get(reverse("api_records", args=["sign"]))
        self.assertEqual(
            [row["ST_ID"] for row in view_site_records.json()["records"]],
            ["A"],
        )

    def test_admin_view_all_records_is_grouped_by_project(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("admin:inventory_tabrecord_changelist"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-project-groups="2"')
        self.assertContains(response, "Alpha")
        self.assertContains(response, "Beta")
        self.assertContains(response, "Sign Inventory: 1", count=2)
        self.assertEqual(len(response.context["project_groups"]), 2)

        filtered = self.client.get(
            reverse("admin:inventory_tabrecord_changelist"),
            {"q": "other-project-user"},
        )
        self.assertContains(filtered, 'data-project-groups="1"')
        self.assertContains(filtered, "Beta")
        self.assertNotContains(filtered, "Alpha")

    def test_project_admin_loads_individual_member_removal_control(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("admin:inventory_project_change", args=[self.alpha.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "inventory/js/project-members")
        self.assertContains(response, 'id="id_members"')

    def test_admin_pending_notification_count_and_approval(self):
        pending_user = User.objects.create_user("waiting-user", "waiting@example.com", "StrongPass!234", is_active=False)
        RegistrationApproval.objects.create(user=pending_user)
        self.client.force_login(self.admin)
        count = self.client.get(reverse("admin:auth_user_pending_approval_count"))
        self.assertEqual(count.json()["count"], 1)
        queue = self.client.get(reverse("admin:auth_user_pending_approvals"))
        self.assertContains(queue, "waiting-user")
        with override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend"):
            response = self.client.post(reverse("admin:auth_user_review_registration", args=[pending_user.pk]), {"action": "approve"})
        self.assertRedirects(response, reverse("admin:auth_user_pending_approvals"))
        pending_user.refresh_from_db()
        self.assertTrue(pending_user.is_active)
        self.assertEqual(pending_user.registration_approval.status, RegistrationApproval.STATUS_APPROVED)

    @override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
    def test_admin_rejects_registration_with_email_reason(self):
        pending_user = User.objects.create_user("rejected-user", "rejected@example.com", "StrongPass!234", is_active=False)
        RegistrationApproval.objects.create(user=pending_user)
        self.client.force_login(self.admin)
        self.client.post(reverse("admin:auth_user_review_registration", args=[pending_user.pk]), {"action": "reject", "reason": "Account could not be verified."})
        approval = RegistrationApproval.objects.get(user=pending_user)
        self.assertEqual(approval.status, RegistrationApproval.STATUS_REJECTED)
        self.assertEqual(approval.rejection_reason, "Account could not be verified.")
        self.assertIn("Account could not be verified.", mail.outbox[-1].body)

    def test_new_record_is_tagged_with_selected_project(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse("api_records", args=["sign"]),
            data=json.dumps({"row": {"ST_ID": "NEW", "POLE_ID": "P3", "SIGN": "S3"}}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(TabRecord.objects.filter(project=self.alpha, data__ST_ID="NEW").exists())

    def test_user_without_project_sees_contact_admin_page(self):
        unassigned = User.objects.create_user("unassigned", password="StrongPass!234")
        self.client.force_login(unassigned)
        response = self.client.get(reverse("home"))
        self.assertContains(response, "Contact Admin")
        api_response = self.client.get(reverse("api_records", args=["sign"]))
        self.assertEqual(api_response.status_code, 403)
