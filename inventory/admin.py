import io
import json
from datetime import date, datetime, timedelta
from zipfile import BadZipFile

from django.contrib import admin
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django.contrib.auth.forms import ReadOnlyPasswordHashWidget, UserChangeForm
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Max
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .forms import ConfigurationExcelImportForm, InventoryExcelImportForm
from .concurrency import section_write_locks
from .excel_security import safe_excel_cell
from .models import (
    AutoFillMapping,
    DropdownOption,
    InventorySection,
    MutcdClassification,
    MutcdFallback,
    MutcdMapping,
    Project,
    RegistrationApproval,
    TabRecord,
)
from .specs import TAB_ORDER, compute_auto_fields, get_section_state, get_spec, missing_required_fields
from .signals import send_account_approved_email, send_account_deactivated_email, send_account_rejected_email

admin.site.site_header = "Bluedome Inventory"
admin.site.site_title = "Bluedome Inventory"
admin.site.index_title = "Inventory Configuration"
admin.site.site_url = "/"


@admin.register(Project)
class ProjectAdmin(admin.ModelAdmin):
    list_display = (
        "project_name",
        "code_badge",
        "status_badge",
        "member_count",
        "created_on",
        "project_actions",
    )
    list_display_links = ("project_name",)
    list_filter = ("is_active",)
    search_fields = ("name", "code", "members__username", "members__email")
    filter_horizontal = ("members",)
    prepopulated_fields = {"code": ("name",)}
    fieldsets = (
        (
            "Project details",
            {
                "fields": ("name", "code", "is_active"),
                "description": (
                    "Create a clear project identity. The project code is generated "
                    "from the name and can be adjusted before saving."
                ),
            },
        ),
        (
            "Project access",
            {
                "fields": ("members",),
                "description": (
                    "Choose the users who may view and manage this project's inventory. "
                    "Use the search boxes to quickly find an account."
                ),
            },
        ),
    )

    class Media:
        js = ("inventory/js/project-members.js",)

    def change_view(self, request, object_id, form_url="", extra_context=None):
        """Keep Admin's project choice in sync with the View Site workspace.

        The inventory workspace and admin use the same session project key. An
        admin opening a project here therefore expects the global "View site"
        link to show that project's records, including rows created by users.
        """
        project = self.get_queryset(request).filter(pk=object_id, is_active=True).first()
        if project is not None:
            request.session["inventory_project_id"] = project.pk
        return super().change_view(request, object_id, form_url, extra_context)

    @admin.display(description="Assigned users")
    def member_count(self, obj):
        count = obj.members.count()
        return format_html(
            '<span class="project-member-count"><span aria-hidden="true">&#128101;</span> {} {}</span>',
            count,
            "user" if count == 1 else "users",
        )

    @admin.display(description="Project", ordering="name")
    def project_name(self, obj):
        return format_html(
            '<span class="project-name-cell"><span class="project-avatar" aria-hidden="true">{}</span>'
            '<span><strong>{}</strong><small>Inventory workspace</small></span></span>',
            obj.name[:1].upper(),
            obj.name,
        )

    @admin.display(description="Code", ordering="code")
    def code_badge(self, obj):
        return format_html('<span class="project-code-badge">{}</span>', obj.code)

    @admin.display(description="Status", ordering="is_active")
    def status_badge(self, obj):
        label = "Active" if obj.is_active else "Inactive"
        state = "active" if obj.is_active else "inactive"
        return format_html(
            '<span class="project-status project-status--{}"><span aria-hidden="true"></span>{}</span>',
            state,
            label,
        )

    @admin.display(description="Created", ordering="created_at")
    def created_on(self, obj):
        return timezone.localtime(obj.created_at).strftime("%m-%d-%Y")

    @admin.display(description="Actions")
    def project_actions(self, obj):
        change_url = reverse("admin:inventory_project_change", args=[obj.pk])
        delete_url = reverse("admin:inventory_project_delete", args=[obj.pk])
        return format_html(
            '<span class="project-row-actions">'
            '<a href="{}" class="project-icon-action project-icon-action--edit" title="Edit project" aria-label="Edit {}">'
            '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M3 17.25V21h3.75L17.81 9.94l-3.75-3.75L3 17.25zM20.71 7.04a1.003 1.003 0 0 0 0-1.42l-2.34-2.34a1.003 1.003 0 0 0-1.42 0l-1.83 1.83 3.75 3.75 1.84-1.82z"/></svg></a>'
            '<a href="{}" class="project-icon-action project-icon-action--delete" title="Delete project" aria-label="Delete {}">'
            '<svg viewBox="0 0 24 24" aria-hidden="true"><path d="M6 19c0 1.1.9 2 2 2h8c1.1 0 2-.9 2-2V7H6v12zm3.46-7.12 1.41-1.41L12 11.59l1.12-1.12 1.41 1.41L13.41 13l1.12 1.12-1.41 1.41L12 14.41l-1.12 1.12-1.41-1.41L10.59 13l-1.13-1.12zM15.5 4l-1-1h-5l-1 1H5v2h14V4z"/></svg></a>'
            '</span>',
            change_url, obj.name, delete_url, obj.name,
        )


User = get_user_model()
try:
    admin.site.unregister(User)
except admin.sites.NotRegistered:
    pass


class BluedomePasswordHashWidget(ReadOnlyPasswordHashWidget):
    template_name = "inventory/widgets/read_only_password_hash.html"


class BluedomeUserChangeForm(UserChangeForm):
    class Meta(UserChangeForm.Meta):
        model = User

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["password"].widget = BluedomePasswordHashWidget()


@admin.register(User)
class BluedomeUserAdmin(DjangoUserAdmin):
    form = BluedomeUserChangeForm
    list_display = (
        "username",
        "email",
        "first_name",
        "last_name",
        "approval_status",
        "is_staff",
        "date_joined",
        "account_actions",
    )
    list_filter = ("is_active", "is_staff", "is_superuser", "date_joined")
    actions = ("approve_selected_users", "deactivate_selected_users")

    @admin.display(description="Account status", ordering="is_active")
    def approval_status(self, obj):
        return "Active" if obj.is_active else "Pending / inactive"

    @admin.display(description="Actions")
    def account_actions(self, obj):
        if obj.is_active:
            if obj.is_superuser:
                return format_html('<span class="account-action-disabled">Protected</span>')
            url = reverse("admin:auth_user_deactivate_account", args=[obj.pk])
            return format_html(
                '<a class="account-action account-action-deactivate" href="{}" '
                'aria-label="Deactivate {}">Deactivate</a>',
                url,
                obj.username,
            )
        url = reverse("admin:auth_user_activate_account", args=[obj.pk])
        return format_html(
            '<a class="account-action account-action-activate" href="{}" '
            'aria-label="Activate {}">Activate</a>',
            url,
            obj.username,
        )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("pending-approvals/", self.admin_site.admin_view(self.pending_approvals), name="auth_user_pending_approvals"),
            path("pending-approvals/count/", self.admin_site.admin_view(self.pending_approval_count), name="auth_user_pending_approval_count"),
            path("<int:user_id>/review-registration/", self.admin_site.admin_view(self.review_registration), name="auth_user_review_registration"),
            path(
                "<int:user_id>/activate-account/",
                self.admin_site.admin_view(self.change_account_status),
                {"activate": True},
                name="auth_user_activate_account",
            ),
            path(
                "<int:user_id>/deactivate-account/",
                self.admin_site.admin_view(self.change_account_status),
                {"activate": False},
                name="auth_user_deactivate_account",
            ),
        ]
        return custom_urls + urls

    def pending_approval_count(self, request):
        return JsonResponse({"count": RegistrationApproval.objects.filter(status=RegistrationApproval.STATUS_PENDING).count()})

    def pending_approvals(self, request):
        pending = RegistrationApproval.objects.filter(status=RegistrationApproval.STATUS_PENDING).select_related("user")
        return render(request, "admin/auth/user/pending_approvals.html", {
            **self.admin_site.each_context(request), "title": "Pending user approvals", "pending_approvals": pending, "opts": self.model._meta,
        })

    def review_registration(self, request, user_id):
        approval = get_object_or_404(RegistrationApproval.objects.select_related("user"), user_id=user_id, status=RegistrationApproval.STATUS_PENDING)
        if request.method == "POST":
            action = request.POST.get("action")
            if action == "approve":
                approval.user._skip_approval_email = True
                approval.user.is_active = True
                approval.user.save(update_fields=["is_active"])
                approval.status = RegistrationApproval.STATUS_APPROVED
                approval.rejection_reason = ""
                approval.reviewed_by = request.user
                approval.reviewed_at = timezone.now()
                approval.save(update_fields=["status", "rejection_reason", "reviewed_by", "reviewed_at"])
                delivered, detail = send_account_approved_email(approval.user_id)
                messages.success(request, f'User "{approval.user.username}" was approved.')
            elif action == "reject":
                reason = request.POST.get("reason", "").strip()
                approval.user.is_active = False
                approval.user.save(update_fields=["is_active"])
                approval.status = RegistrationApproval.STATUS_REJECTED
                approval.rejection_reason = reason
                approval.reviewed_by = request.user
                approval.reviewed_at = timezone.now()
                approval.save(update_fields=["status", "rejection_reason", "reviewed_by", "reviewed_at"])
                delivered, detail = send_account_rejected_email(approval.user_id, reason)
                messages.success(request, f'User "{approval.user.username}" was rejected.')
            else:
                messages.error(request, "Choose Approve or Reject.")
                return redirect("admin:auth_user_review_registration", user_id=user_id)
            messages.success(request, detail) if delivered else messages.warning(request, detail)
            return redirect("admin:auth_user_pending_approvals")
        return render(request, "admin/auth/user/review_registration.html", {
            **self.admin_site.each_context(request), "title": "Review registration", "approval": approval, "opts": self.model._meta,
        })

    def change_account_status(self, request, user_id, activate):
        if not self.has_change_permission(request):
            raise PermissionDenied
        user = get_object_or_404(User, pk=user_id)
        if not activate and (user.is_superuser or user.pk == request.user.pk):
            messages.error(request, "The current administrator or a superuser cannot be deactivated here.")
            return redirect("admin:auth_user_changelist")
        if request.method == "POST":
            user._skip_approval_email = activate
            user._skip_deactivation_email = not activate
            user.is_active = activate
            user.save(update_fields=["is_active"])
            if activate:
                RegistrationApproval.objects.filter(user=user, status=RegistrationApproval.STATUS_PENDING).update(status=RegistrationApproval.STATUS_APPROVED, reviewed_by=request.user, reviewed_at=timezone.now())
            action = "activated" if activate else "deactivated"
            messages.success(request, f'User "{user.username}" was {action}.')
            delivery_function = (
                send_account_approved_email if activate else send_account_deactivated_email
            )
            delivered, delivery_message = delivery_function(user.pk)
            if delivered:
                messages.success(request, delivery_message)
            else:
                messages.warning(request, delivery_message)
            return redirect("admin:auth_user_changelist")
        return render(
            request,
            "admin/auth/user/account_status_confirmation.html",
            {
                **self.admin_site.each_context(request),
                "title": f'{"Activate" if activate else "Deactivate"} user',
                "target_user": user,
                "activate": activate,
                "opts": self.model._meta,
            },
        )

    @admin.action(description="Approve selected users and send activation email")
    def approve_selected_users(self, request, queryset):
        approved = 0
        delivered = 0
        failures = []
        for user in queryset.filter(is_active=False):
            user._skip_approval_email = True
            user.is_active = True
            user.save(update_fields=["is_active"])
            RegistrationApproval.objects.filter(
                user=user, status=RegistrationApproval.STATUS_PENDING
            ).update(
                status=RegistrationApproval.STATUS_APPROVED,
                reviewed_by=request.user,
                reviewed_at=timezone.now(),
            )
            approved += 1
            was_delivered, delivery_message = send_account_approved_email(user.pk)
            if was_delivered:
                delivered += 1
            else:
                failures.append(f"{user.username}: {delivery_message}")
        self.message_user(request, f"Approved {approved} user account(s).")
        if delivered:
            self.message_user(request, f"Sent {delivered} approval email(s).", level=messages.SUCCESS)
        if failures:
            self.message_user(request, "Email not delivered — " + "; ".join(failures), level=messages.WARNING)

    @admin.action(description="Deactivate selected users")
    def deactivate_selected_users(self, request, queryset):
        eligible = queryset.filter(is_active=True, is_superuser=False).exclude(pk=request.user.pk)
        deactivated = 0
        delivered = 0
        failures = []
        for user in eligible:
            user._skip_deactivation_email = True
            user.is_active = False
            user.save(update_fields=["is_active"])
            deactivated += 1
            was_delivered, delivery_message = send_account_deactivated_email(user.pk)
            if was_delivered:
                delivered += 1
            else:
                failures.append(f"{user.username}: {delivery_message}")
        self.message_user(request, f"Deactivated {deactivated} user account(s).")
        if delivered:
            self.message_user(request, f"Sent {delivered} deactivation email(s).", level=messages.SUCCESS)
        if failures:
            self.message_user(request, "Email not delivered — " + "; ".join(failures), level=messages.WARNING)


class RowActionsAdminMixin:
    """Expose explicit, permission-aware row actions in each change list."""

    @admin.display(description="Actions")
    def row_actions(self, obj):
        opts = obj._meta
        change_url = reverse(
            f"admin:{opts.app_label}_{opts.model_name}_change",
            args=[obj.pk],
        )
        delete_url = reverse(
            f"admin:{opts.app_label}_{opts.model_name}_delete",
            args=[obj.pk],
        )
        return format_html(
            '<a class="row-action row-action-edit" href="{}" '
            'title="Edit" aria-label="Edit">'
            '<svg aria-hidden="true" viewBox="0 0 24 24">'
            '<path d="M4 16.5V20h3.5L18 9.5 14.5 6 4 16.5zm16.7-9.8a1 1 0 0 0 '
            '0-1.4l-2-2a1 1 0 0 0-1.4 0L15.5 5l3.5 3.5 1.7-1.8z"/></svg>'
            '<span class="sr-only">Edit</span></a>'
            '<a class="row-action row-action-delete" href="{}" '
            'title="Delete" aria-label="Delete">'
            '<svg aria-hidden="true" viewBox="0 0 24 24">'
            '<path d="M6 19a2 2 0 0 0 2 2h8a2 2 0 0 0 2-2V7H6v12zM8 '
            '9h8v10H8V9zm7.5-5-1-1h-5l-1 1H5v2h14V4z"/></svg>'
            '<span class="sr-only">Delete</span></a>',
            change_url,
            delete_url,
        )


class ConfigurationExcelAdminMixin:
    """Bulk Excel import/export for normalized admin configuration tables."""

    change_list_template = "admin/inventory/configuration_change_list.html"
    change_form_template = "admin/inventory/configuration_change_form.html"
    excel_columns = ()
    excel_unique_fields = ()
    excel_json_fields = ()

    def sync_related_import_rows(self, pending):
        """Hook for models whose imported rows maintain related lookup data."""

    def get_urls(self):
        opts = self.model._meta
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel),
                name=f"{opts.app_label}_{opts.model_name}_import_excel",
            ),
            path(
                "export-excel/",
                self.admin_site.admin_view(self.export_excel),
                name=f"{opts.app_label}_{opts.model_name}_export_excel",
            ),
            path(
                "excel-template/",
                self.admin_site.admin_view(self.excel_template),
                name=f"{opts.app_label}_{opts.model_name}_excel_template",
            ),
        ]
        return custom_urls + urls

    def _excel_url(self, action):
        opts = self.model._meta
        suffix = {
            "import": "import_excel",
            "export": "export_excel",
            "template": "excel_template",
        }[action]
        return reverse(f"admin:{opts.app_label}_{opts.model_name}_{suffix}")

    def changelist_view(self, request, extra_context=None):
        extra_context = {
            **(extra_context or {}),
            "excel_import_url": self._excel_url("import"),
            "excel_export_url": self._excel_url("export"),
            "excel_template_url": self._excel_url("template"),
        }
        return super().changelist_view(request, extra_context)

    def changeform_view(self, request, object_id=None, form_url="", extra_context=None):
        extra_context = {
            **(extra_context or {}),
            "excel_import_url": self._excel_url("import"),
            "excel_export_url": self._excel_url("export"),
            "excel_template_url": self._excel_url("template"),
        }
        return super().changeform_view(
            request,
            object_id=object_id,
            form_url=form_url,
            extra_context=extra_context,
        )

    @staticmethod
    def _cell_value(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _workbook(self, include_records):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.model._meta.verbose_name_plural.title()[:31]
        headers = [label for label, _field in self.excel_columns]
        TabRecordAdmin._style_sheet(worksheet, headers)
        if include_records:
            for obj in self.get_queryset(None).order_by(*self.model._meta.ordering):
                row = []
                for _label, field_name in self.excel_columns:
                    if field_name == "section":
                        value = obj.section_id
                    else:
                        value = getattr(obj, field_name)
                    if field_name in self.excel_json_fields:
                        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
                    row.append(safe_excel_cell(value))
                worksheet.append(row)
        return workbook

    def _workbook_response(self, workbook, filename):
        output = io.BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def excel_template(self, request):
        if not self.has_view_permission(request):
            raise PermissionDenied
        return self._workbook_response(
            self._workbook(include_records=False),
            f"{self.model._meta.model_name}-import-template.xlsx",
        )

    def export_excel(self, request):
        if not self.has_view_permission(request):
            raise PermissionDenied
        return self._workbook_response(
            self._workbook(include_records=True),
            f"{self.model._meta.model_name}-all-records.xlsx",
        )

    def import_excel(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied
        opts = self.model._meta
        if request.method == "POST":
            form = ConfigurationExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    workbook = load_workbook(
                        form.cleaned_data["excel_file"],
                        read_only=True,
                        data_only=True,
                    )
                except (BadZipFile, InvalidFileException, OSError, ValueError):
                    form.add_error("excel_file", "The uploaded file is not a valid Excel workbook.")
                else:
                    worksheet = workbook.active
                    expected_headers = [label for label, _field in self.excel_columns]
                    headers = [
                        self._cell_value(value)
                        for value in next(
                            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                            (),
                        )
                    ]
                    while headers and not headers[-1]:
                        headers.pop()
                    errors = []
                    if headers != expected_headers:
                        errors.append(
                            "The headers must exactly match the downloaded template: "
                            + ", ".join(expected_headers)
                        )
                    pending = []
                    seen = set()
                    if not errors:
                        for row_number, values in enumerate(
                            worksheet.iter_rows(min_row=2, values_only=True),
                            start=2,
                        ):
                            if row_number > 10001:
                                errors.append("A maximum of 10,000 rows can be imported at once.")
                                break
                            if not any(value not in (None, "") for value in values):
                                continue
                            data = {}
                            for index, (_label, field_name) in enumerate(self.excel_columns):
                                value = self._cell_value(
                                    values[index] if index < len(values) else ""
                                )
                                if field_name == "section":
                                    data["section_id"] = value.lower()
                                elif field_name in self.excel_json_fields:
                                    try:
                                        data[field_name] = json.loads(value or "{}")
                                    except json.JSONDecodeError:
                                        errors.append(
                                            f"Row {row_number}: {field_name} must contain valid JSON."
                                        )
                                elif self.model._meta.get_field(field_name).get_internal_type() in {
                                    "IntegerField",
                                    "PositiveIntegerField",
                                }:
                                    try:
                                        data[field_name] = int(value or 0)
                                    except ValueError:
                                        errors.append(
                                            f"Row {row_number}: {field_name} must be a whole number."
                                        )
                                else:
                                    data[field_name] = value
                            unique_key = tuple(
                                data.get("section_id" if field == "section" else field)
                                for field in self.excel_unique_fields
                            )
                            if unique_key in seen:
                                errors.append(f"Row {row_number}: duplicate record in workbook.")
                            seen.add(unique_key)
                            obj = self.model(**data)
                            try:
                                # Existing unique keys are valid during import: they are
                                # updated atomically below. Field/FK validation still runs.
                                obj.full_clean(
                                    validate_unique=False,
                                    validate_constraints=False,
                                )
                            except ValidationError as exc:
                                details = "; ".join(
                                    f"{field}: {', '.join(messages)}"
                                    for field, messages in exc.message_dict.items()
                                )
                                errors.append(f"Row {row_number}: {details}")
                            pending.append(obj)
                    if errors:
                        form.add_error(None, errors[:25])
                    elif not pending:
                        form.add_error(None, "The workbook does not contain any data rows.")
                    else:
                        try:
                            with transaction.atomic():
                                created_count = 0
                                updated_count = 0
                                for obj in pending:
                                    lookup = {
                                        "section_id" if field == "section" else field: getattr(
                                            obj,
                                            "section_id" if field == "section" else field,
                                        )
                                        for field in self.excel_unique_fields
                                    }
                                    defaults = {
                                        field.name: getattr(obj, field.name)
                                        for field in self.model._meta.concrete_fields
                                        if not field.primary_key
                                        and field.name not in self.excel_unique_fields
                                        and field.name != "section"
                                    }
                                    _saved, created = self.model.objects.update_or_create(
                                        defaults=defaults,
                                        **lookup,
                                    )
                                    if created:
                                        created_count += 1
                                    else:
                                        updated_count += 1
                                self.sync_related_import_rows(pending)
                        except IntegrityError:
                            form.add_error(
                                None,
                                "The import conflicts with data saved by another request. "
                                "Export the latest data and try again.",
                            )
                        else:
                            messages.success(
                                request,
                                f"Successfully imported {len(pending)} "
                                f"{opts.verbose_name_plural}: {created_count} created, "
                                f"{updated_count} updated.",
                            )
                            return redirect(
                                reverse(
                                    f"admin:{opts.app_label}_{opts.model_name}_changelist"
                                )
                            )
        else:
            form = ConfigurationExcelImportForm()
        return render(
            request,
            "admin/inventory/configuration_import_excel.html",
            {
                **self.admin_site.each_context(request),
                "form": form,
                "title": f"Import {opts.verbose_name_plural}",
                "opts": opts,
                "changelist_url": reverse(
                    f"admin:{opts.app_label}_{opts.model_name}_changelist"
                ),
                "template_url": self._excel_url("template"),
                "export_url": self._excel_url("export"),
            },
        )
@admin.register(InventorySection)
class InventorySectionAdmin(admin.ModelAdmin):
    list_display = ("name", "key", "updated_at")
    search_fields = ("name", "key")
    readonly_fields = ("key", "updated_at")

    def has_add_permission(self, request):
        # The four supported sections are created by the seed migration.
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(DropdownOption)
class DropdownOptionAdmin(ConfigurationExcelAdminMixin, RowActionsAdminMixin, admin.ModelAdmin):
    excel_columns = (
        ("Section", "section"),
        ("Field name", "field_name"),
        ("Value", "value"),
        ("Sort order", "sort_order"),
    )
    excel_unique_fields = ("section", "field_name", "value")
    list_display = ("section", "field_name", "value", "sort_order", "row_actions")
    list_filter = ("section", "field_name")
    search_fields = ("field_name", "value")
    ordering = ("section", "field_name", "sort_order", "id")


@admin.register(AutoFillMapping)
class AutoFillMappingAdmin(ConfigurationExcelAdminMixin, RowActionsAdminMixin, admin.ModelAdmin):
    excel_columns = (
        ("Section", "section"),
        ("Driver value", "driver_value"),
        ("Values JSON", "values"),
    )
    excel_unique_fields = ("section", "driver_value")
    excel_json_fields = ("values",)
    list_display = ("section", "driver_value", "row_actions")
    list_filter = ("section",)
    search_fields = ("driver_value",)


@admin.register(MutcdMapping)
class MutcdMappingAdmin(ConfigurationExcelAdminMixin, RowActionsAdminMixin, admin.ModelAdmin):
    excel_columns = (
        ("Section", "section"),
        ("Word description", "word_description"),
        ("MUTCD code", "mutcd_code"),
        ("Classification", "classification"),
    )
    excel_unique_fields = ("section", "word_description")
    list_display = ("word_description", "mutcd_code", "classification", "section", "row_actions")
    list_filter = ("section", "classification")
    search_fields = ("word_description", "mutcd_code", "classification")

    def sync_related_import_rows(self, pending):
        """Keep MUTCD classification and one deterministic fallback in sync."""
        related_by_code = {}
        for mapping in pending:
            if mapping.mutcd_code:
                related_by_code.setdefault(
                    (mapping.section_id, mapping.mutcd_code),
                    (mapping.word_description, mapping.classification),
                )

        for (section_id, mutcd_code), (word_description, classification) in related_by_code.items():
            MutcdFallback.objects.update_or_create(
                section_id=section_id,
                code=mutcd_code,
                defaults={"word_description": word_description},
            )
            if classification:
                MutcdClassification.objects.update_or_create(
                    section_id=section_id,
                    code=mutcd_code,
                    defaults={"classification": classification},
                )


@admin.register(MutcdClassification)
class MutcdClassificationAdmin(
    ConfigurationExcelAdminMixin,
    RowActionsAdminMixin,
    admin.ModelAdmin,
):
    excel_columns = (
        ("Section", "section"),
        ("MUTCD code", "code"),
        ("Classification", "classification"),
    )
    excel_unique_fields = ("section", "code")
    list_display = ("code", "classification", "section", "row_actions")
    list_filter = ("section", "classification")
    search_fields = ("code", "classification")


@admin.register(MutcdFallback)
class MutcdFallbackAdmin(ConfigurationExcelAdminMixin, RowActionsAdminMixin, admin.ModelAdmin):
    excel_columns = (
        ("Section", "section"),
        ("MUTCD code", "code"),
        ("Word description", "word_description"),
    )
    excel_unique_fields = ("section", "code")
    list_display = ("code", "word_description", "section", "row_actions")
    list_filter = ("section",)
    search_fields = ("code", "word_description")


class InventoryUsernameFilter(admin.SimpleListFilter):
    title = "Username"
    parameter_name = "username"

    def lookups(self, request, model_admin):
        owner_ids = (
            TabRecord.objects.exclude(owner_id=None)
            .order_by()
            .values_list("owner_id", flat=True)
            .distinct()
        )
        users = get_user_model().objects.filter(pk__in=owner_ids).order_by("username")
        choices = [
            (str(user.pk), user.get_full_name().strip() or user.username)
            for user in users
        ]
        if TabRecord.objects.filter(owner_id=None).exists():
            choices.append(("legacy", "Legacy / unknown"))
        return choices

    def queryset(self, request, queryset):
        if self.value() == "legacy":
            return queryset.filter(owner_id=None)
        if self.value():
            return queryset.filter(owner_id=self.value())
        return queryset


class InventorySectionFilter(admin.SimpleListFilter):
    title = "Inventory section"
    parameter_name = "inventory_section"

    def lookups(self, request, model_admin):
        tabs = list(
            TabRecord.objects.order_by().values_list("tab", flat=True).distinct()
        )
        section_names = dict(
            InventorySection.objects.filter(key__in=tabs).values_list("key", "name")
        )
        return [
            (tab, section_names.get(tab, tab.replace("_", " ").title()))
            for tab in sorted(tabs, key=str.casefold)
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(tab=self.value())
        return queryset


class InventoryDateAddedFilter(admin.ListFilter):
    title = "Date added"
    template = "admin/inventory/tabrecord/date_range_filter.html"

    def __init__(self, request, params, model, model_admin):
        super().__init__(request, params, model, model_admin)
        self.date_from = params.pop("date_from", [""])[-1]
        self.date_to = params.pop("date_to", [""])[-1]
        self.other_parameters = [
            (key, value)
            for key, values in request.GET.lists()
            if key not in {"date_from", "date_to", "p"}
            for value in values
        ]

    def has_output(self):
        return True

    def expected_parameters(self):
        return ["date_from", "date_to"]

    def choices(self, changelist):
        return ()

    @staticmethod
    def _parse(value):
        try:
            return datetime.strptime(value, "%m-%d-%Y").date()
        except (TypeError, ValueError):
            return None

    def queryset(self, request, queryset):
        start_date = self._parse(self.date_from)
        end_date = self._parse(self.date_to)
        if start_date:
            start = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
            queryset = queryset.filter(created_at__gte=start)
        if end_date:
            end = timezone.make_aware(
                datetime.combine(end_date + timedelta(days=1), datetime.min.time())
            )
            queryset = queryset.filter(created_at__lt=end)
        return queryset


@admin.register(TabRecord)
class TabRecordAdmin(admin.ModelAdmin):
    list_display = ("record_id", "username", "tab", "date_added")
    list_filter = (
        InventoryUsernameFilter,
        InventorySectionFilter,
        InventoryDateAddedFilter,
    )
    search_fields = ("tab_record_id", "owner__username", "owner__email")
    ordering = ("tab", "tab_record_id")
    change_list_template = "admin/inventory/tabrecord/change_list.html"

    @staticmethod
    def _selected_project(request):
        project_id = request.session.get("inventory_project_id")
        project = Project.objects.filter(pk=project_id, is_active=True).first()
        if project is None:
            project = Project.objects.filter(is_active=True).first()
            if project:
                request.session["inventory_project_id"] = project.pk
        return project

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("project", "owner")

    def changelist_view(self, request, extra_context=None):
        """Render the filtered/paginated admin results grouped by project."""
        try:
            changelist = self.get_changelist_instance(request)
            records = list(changelist.result_list)
        except Exception:
            # Let Django render its normal validation/error response for an
            # invalid query-string filter rather than masking the exception.
            return super().changelist_view(request, extra_context=extra_context)

        grouped = {}
        for record in records:
            group = grouped.setdefault(
                record.project_id,
                {
                    "project": record.project,
                    "records": [],
                    "section_counts": {},
                },
            )
            group["records"].append(record)
            label = get_spec(record.tab)["tab_label"]
            group["section_counts"][label] = group["section_counts"].get(label, 0) + 1

        project_groups = sorted(grouped.values(), key=lambda group: group["project"].name.casefold())
        for index, group in enumerate(project_groups):
            group["section_counts"] = sorted(group["section_counts"].items())
            group["expanded"] = len(project_groups) <= 3 or index == 0

        context = {
            "project_groups": project_groups,
            "grouped_record_count": len(records),
            **(extra_context or {}),
        }
        return super().changelist_view(request, extra_context=context)

    @admin.display(description="ID", ordering="tab_record_id")
    def record_id(self, obj):
        return obj.tab_record_id

    @admin.display(description="Username", ordering="owner__username")
    def username(self, obj):
        if not obj.owner:
            return "Legacy / unknown"
        return obj.owner.get_full_name().strip() or obj.owner.username

    @admin.display(description="Date", ordering="created_at")
    def date_added(self, obj):
        return timezone.localtime(obj.created_at).strftime("%m-%d-%Y")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel),
                name="inventory_tabrecord_import_excel",
            ),
            path(
                "export-excel/",
                self.admin_site.admin_view(self.export_excel),
                name="inventory_tabrecord_export_excel",
            ),
            path(
                "excel-template/",
                self.admin_site.admin_view(self.excel_template),
                name="inventory_tabrecord_excel_template",
            ),
        ]
        return custom_urls + urls

    @staticmethod
    def _excel_value(value):
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%m-%d-%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _style_sheet(ws, columns):
        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(bold=True, color="FFFFFF")
        for index, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=index, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(index)].width = max(14, min(35, len(column) + 4))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    def import_excel(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied
        project = self._selected_project(request)
        if project is None:
            messages.warning(request, "Select an active project from the main inventory page first.")
            return redirect("home")

        if request.method == "POST":
            form = InventoryExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                section_key = form.cleaned_data["section"].key
                spec = get_spec(section_key)
                try:
                    workbook = load_workbook(
                        form.cleaned_data["excel_file"],
                        read_only=True,
                        data_only=True,
                    )
                except (BadZipFile, InvalidFileException, OSError, ValueError):
                    form.add_error("excel_file", "The uploaded file is not a valid Excel workbook.")
                else:
                    sheet_name = spec["export_sheet_name"]
                    worksheet = (
                        workbook[sheet_name]
                        if sheet_name in workbook.sheetnames
                        else workbook.active
                    )
                    raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                    headers = [self._excel_value(value) for value in raw_headers]
                    while headers and not headers[-1]:
                        headers.pop()
                    duplicate_headers = {header for header in headers if header and headers.count(header) > 1}
                    unknown_headers = [header for header in headers if header and header not in spec["columns"]]
                    missing_uid_headers = [field for field in spec["uid_parts"] if field not in headers]

                    errors = []
                    if not headers:
                        errors.append("The workbook does not contain a header row.")
                    if duplicate_headers:
                        errors.append("Duplicate headers: " + ", ".join(sorted(duplicate_headers)))
                    if unknown_headers:
                        errors.append("Unknown headers: " + ", ".join(unknown_headers))
                    if missing_uid_headers:
                        errors.append("Required headers are missing: " + ", ".join(missing_uid_headers))

                    pending_rows = []
                    state = get_section_state(section_key)
                    if not errors:
                        for excel_row_number, values in enumerate(
                            worksheet.iter_rows(min_row=2, values_only=True),
                            start=2,
                        ):
                            if excel_row_number > 10001:
                                errors.append("A maximum of 10,000 data rows can be imported at once.")
                                break
                            if not any(value not in (None, "") for value in values):
                                continue
                            supplied = {
                                header: self._excel_value(values[index] if index < len(values) else "")
                                for index, header in enumerate(headers)
                                if header and header not in {"ID", "IMAGE_LINK"}
                            }
                            row = {
                                column: supplied.get(column, "")
                                for column in spec["columns"]
                                if column != "ID"
                            }
                            row = compute_auto_fields(section_key, row, state)
                            row["IMAGE_LINK"] = ""
                            missing = missing_required_fields(section_key, row)
                            if missing:
                                errors.append(
                                    f"Row {excel_row_number}: missing required values "
                                    + ", ".join(missing)
                                )
                            pending_rows.append(row)

                    if errors:
                        form.add_error(None, errors[:25])
                    elif not pending_rows:
                        form.add_error(None, "The workbook does not contain any data rows.")
                    else:
                        try:
                            with section_write_locks[section_key]:
                                with transaction.atomic():
                                    max_id = (
                                        TabRecord.objects.filter(tab=section_key)
                                        .aggregate(value=Max("tab_record_id"))["value"]
                                        or 0
                                    )
                                    TabRecord.objects.bulk_create(
                                        [
                                            TabRecord(
                                                project=project,
                                                tab=section_key,
                                                tab_record_id=max_id + offset,
                                                data=row,
                                                owner=request.user,
                                            )
                                            for offset, row in enumerate(pending_rows, start=1)
                                        ]
                                    )
                        except IntegrityError:
                            form.add_error(
                                None,
                                "The import conflicts with inventory records saved by "
                                "another request. Export the latest data and try again.",
                            )
                        else:
                            self.message_user(
                                request,
                                f"Successfully imported {len(pending_rows)} "
                                f"{spec['tab_label']} records.",
                                messages.SUCCESS,
                            )
                            return redirect("admin:inventory_tabrecord_changelist")
        else:
            form = InventoryExcelImportForm()

        context = {
            **self.admin_site.each_context(request),
            "title": "Import inventory records from Excel",
            "form": form,
            "opts": self.model._meta,
        }
        return render(request, "admin/inventory/tabrecord/import_excel.html", context)

    def excel_template(self, request):
        workbook = Workbook()
        workbook.remove(workbook.active)
        for section_key in TAB_ORDER:
            spec = get_spec(section_key)
            ws = workbook.create_sheet(spec["export_sheet_name"])
            self._style_sheet(ws, spec["columns"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Bluedome_Inventory_Import_Template.xlsx"'
        return response

    def export_excel(self, request):
        project = self._selected_project(request)
        if project is None:
            messages.warning(request, "Select an active project from the main inventory page first.")
            return redirect("home")
        workbook = Workbook()
        workbook.remove(workbook.active)
        record_count = 0
        for section_key in TAB_ORDER:
            spec = get_spec(section_key)
            records = list(
                TabRecord.objects.filter(tab=section_key, project=project)
                .select_related("owner")
                .order_by("tab_record_id")
            )
            ws = workbook.create_sheet(spec["export_sheet_name"])
            export_columns = list(spec["columns"]) + ["ADDED_BY"]
            self._style_sheet(ws, export_columns)
            for record in records:
                row = record.as_row(include_owner=True)
                ws.append([safe_excel_cell(row.get(column, "")) for column in export_columns])
            record_count += len(records)
        buffer = io.BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="Bluedome_{project.code}_Inventory_Records_{date.today().isoformat()}.xlsx"'
        )
        response["X-Record-Count"] = str(record_count)
        return response
