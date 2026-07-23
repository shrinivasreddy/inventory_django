import io
import hashlib
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.password_validation import get_default_password_validators
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import FileResponse, JsonResponse, HttpResponseNotAllowed, HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_GET, require_http_methods, require_POST
from django.views.decorators.cache import never_cache
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import (
    AutoFillMapping,
    DropdownOption,
    MutcdClassification,
    MutcdMapping,
    TabRecord,
)
from .forms import SignUpForm
from .assistant import AssistantError, interpret_inventory_prompt, transcribe_audio
from .concurrency import section_write_locks as locks
from .excel_security import safe_excel_cell
from .image_storage import (
    FOLDER_SECTIONS,
    InventoryImageError,
    record_image_directory,
    save_record_image,
)
from .specs import (
    TAB_ORDER,
    compute_auto_fields,
    get_section_state,
    get_spec,
    missing_required_fields,
)

def invalid_date_fields(spec, row):
    """Return date fields that are not valid MM-DD-YYYY calendar dates."""
    invalid = []
    for field in spec["date_fields"]:
        value = row.get(field, "").strip()
        if not value:
            continue
        try:
            datetime.strptime(value, "%m-%d-%Y")
        except ValueError:
            invalid.append(field)
    return invalid


def normalize_coordinate_fields(spec, row):
    """Round coordinates to 7 decimals and validate geographic ranges."""
    errors = []
    precision = Decimal("0.0000001")
    for field in spec["columns"]:
        if not (field.endswith("LATITUDE") or field.endswith("LONGITUDE")):
            continue
        value = row.get(field, "").strip()
        if not value:
            continue
        try:
            coordinate = Decimal(value)
        except InvalidOperation:
            errors.append(f"{field} must be a number")
            continue
        if not coordinate.is_finite():
            errors.append(f"{field} must be a finite number")
            continue
        limit = Decimal("90") if field.endswith("LATITUDE") else Decimal("180")
        if coordinate < -limit or coordinate > limit:
            errors.append(f"{field} must be between {-limit} and {limit}")
            continue
        rounded = coordinate.quantize(precision, rounding=ROUND_HALF_UP)
        row[field] = format(rounded, "f").rstrip("0").rstrip(".") or "0"
    return errors


def api_login_required(view):
    """Return a machine-readable 401 for API clients instead of an HTML
    redirect to the login form."""
    @wraps(view)
    def wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({"error": "Authentication required."}, status=401)
        return view(request, *args, **kwargs)
    return wrapped


def parse_json_body(request):
    try:
        return json.loads(request.body or b"{}")
    except (json.JSONDecodeError, UnicodeDecodeError):
        return {}


def next_tab_id(key):
    from django.db.models import Max
    max_id = TabRecord.objects.filter(tab=key).aggregate(m=Max("tab_record_id"))["m"]
    return (max_id or 0) + 1


def visible_tab_records(request, key):
    queryset = TabRecord.objects.filter(tab=key).select_related("owner")
    if not request.user.is_staff:
        queryset = queryset.filter(owner=request.user)
    return queryset


def renumber_tab_ids(key):
    """Ported from renumber_ids: re-sequence to 1..N after a delete. Safe to
    do as individual .save() calls in ascending order -- each target slot is
    always freed by the previous iteration before it's reused, so the
    (tab, tab_record_id) unique constraint is never violated mid-loop."""
    records = list(TabRecord.objects.filter(tab=key).order_by("tab_record_id"))
    for i, r in enumerate(records, start=1):
        if r.tab_record_id != i:
            r.tab_record_id = i
            r.save(update_fields=["tab_record_id"])


# ---------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------
@require_http_methods(["GET", "POST"])
def signup(request):
    if request.user.is_authenticated:
        return redirect("home")
    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("signup_pending")
    else:
        form = SignUpForm()
    return render(request, "registration/signup.html", {"form": form})


@require_GET
def signup_pending(request):
    return render(request, "registration/signup_pending.html")


@require_POST
def password_requirements(request):
    data = parse_json_body(request)
    password = str(data.get("password") or "")
    user = get_user_model()(
        username=str(data.get("username") or ""),
        email=str(data.get("email") or ""),
    )
    checks = {}
    for validator in get_default_password_validators():
        key = validator.__class__.__name__
        try:
            validator.validate(password, user)
            checks[key] = True
        except ValidationError:
            checks[key] = False
    return JsonResponse({"checks": checks})


@login_required
@require_GET
def home(request):
    tabs = [{"key": k, "label": get_spec(k)["tab_label"]} for k in TAB_ORDER]
    return render(request, "index.html", {"tabs_json": json.dumps(tabs)})


@api_login_required
@require_POST
def api_assistant_preview(request):
    data = parse_json_body(request)
    prompt = str(data.get("prompt") or "").strip()
    if not prompt:
        return JsonResponse({"error": "Enter or speak an instruction first."}, status=400)
    if len(prompt) > 4000:
        return JsonResponse(
            {"error": "The instruction must be 4,000 characters or fewer."},
            status=400,
        )
    safety_identifier = hashlib.sha256(
        f"{request.user.pk}:{request.user.date_joined.isoformat()}".encode("utf-8")
    ).hexdigest()
    try:
        preview = interpret_inventory_prompt(prompt, safety_identifier)
    except AssistantError as exc:
        return JsonResponse({"error": str(exc)}, status=503)

    state = get_section_state(preview["section"])
    spec = get_spec(preview["section"])
    complete_row = {
        column: str(preview["row"].get(column, "") or "")
        for column in spec["columns"]
        if column != "ID"
    }
    complete_row = compute_auto_fields(preview["section"], complete_row, state)
    preview["row"] = complete_row
    preview["missing_required"] = missing_required_fields(
        preview["section"],
        complete_row,
    )
    return JsonResponse(preview)


@api_login_required
@require_POST
def api_assistant_transcribe(request):
    audio = request.FILES.get("audio")
    if audio is None:
        return JsonResponse({"error": "No voice recording was provided."}, status=400)
    if audio.size > 10 * 1024 * 1024:
        return JsonResponse({"error": "Voice recordings must be 10 MB or smaller."}, status=400)
    allowed_audio_types = {
        "audio/webm", "audio/wav", "audio/x-wav", "audio/mpeg",
        "audio/mp4", "audio/ogg", "video/webm",
    }
    if audio.content_type not in allowed_audio_types:
        return JsonResponse({"error": "The uploaded file is not a supported voice recording."}, status=400)
    try:
        text = transcribe_audio(audio.read(), audio.name, audio.content_type)
    except AssistantError as exc:
        return JsonResponse({"error": str(exc)}, status=503)
    return JsonResponse({"text": text})


# ---------------------------------------------------------------------------
# Spec
# ---------------------------------------------------------------------------
@api_login_required
@require_GET
@never_cache
def api_spec(request, key):
    if key not in TAB_ORDER:
        return JsonResponse({"error": "Unknown tab"}, status=404)
    spec = get_spec(key)
    state = get_section_state(key)
    return JsonResponse({
        "key": key,
        "columns": spec["columns"],
        "dropdown_fields": spec["dropdown_fields"],
        "text_fields": spec["text_fields"],
        "auto_fields": spec["auto_fields"],
        "image_field": "IMAGE_LINK",
        "date_fields": spec["date_fields"],
        "uid_field": spec["uid_field"],
        "uid_prefix": spec["uid_prefix"],
        "uid_parts": spec["uid_parts"],
        "uid_required_msg": spec["uid_required_msg"],
        "preserve_case_fields": spec["preserve_case_fields"],
        "searchable_fields": spec.get("searchable_fields", []),
        "wide_cols": spec["wide_cols"],
        "sticky_fields": spec["sticky_fields"],
        "banner_title": spec["banner_title"],
        "banner_subtitle": spec["banner_subtitle"],
        "options": state["options"],
        "mutcd_map": state["mutcd_map"],
        "mutcd_to_class": state["mutcd_to_class"],
        "mutcd_word_options": state["mutcd_word_options"],
        "mutcd_reverse_map": state["mutcd_reverse_map"],
        "auto_fill_map": spec.get("auto_fill_map"),
        "type_map": state["type_map"],
        "conditional_dropdowns": spec.get("conditional_dropdowns", {}),
        "default_field_values": spec.get("default_field_values", {}),
        "can_manage_configuration": request.user.is_staff,
    })


# ---------------------------------------------------------------------------
# Records: GET (list) / POST (add) / DELETE (delete-all)
# ---------------------------------------------------------------------------
@api_login_required
def api_records(request, key):
    if key not in TAB_ORDER:
        return JsonResponse({"error": "Unknown tab"}, status=404)
    spec = get_spec(key)

    if request.method == "GET":
        with locks[key]:
            qs = visible_tab_records(request, key).order_by("tab_record_id")
            records = [r.as_row() for r in qs]
            nid = next_tab_id(key)
        return JsonResponse({"records": records, "next_id": nid})

    if request.method == "POST":
        body = parse_json_body(request)
        raw_row = body.get("row")
        if not isinstance(raw_row, dict):
            return JsonResponse({"error": "Malformed request: 'row' must be an object."}, status=400)
        row = {c: str(raw_row.get(c, "") or "") for c in spec["columns"]}
        row.pop("ID", None)
        row["IMAGE_LINK"] = ""
        invalid_dates = invalid_date_fields(spec, row)
        if invalid_dates:
            return JsonResponse(
                {"error": f"{', '.join(invalid_dates)} must use MM-DD-YYYY."},
                status=400,
            )
        coordinate_errors = normalize_coordinate_fields(spec, row)
        if coordinate_errors:
            return JsonResponse({"error": "; ".join(coordinate_errors)}, status=400)

        with locks[key]:
            ts_dict = get_section_state(key)
            row = compute_auto_fields(key, row, ts_dict)
            missing = missing_required_fields(key, row)
            if missing:
                return JsonResponse({"error": spec["uid_required_msg"]}, status=400)
            try:
                with transaction.atomic():
                    new_id = next_tab_id(key)
                    rec = TabRecord.objects.create(
                        tab=key,
                        tab_record_id=new_id,
                        data=row,
                        owner=request.user,
                    )
            except Exception:
                return JsonResponse({"error": "Failed to save the record to the database."}, status=500)
            return JsonResponse({
                "record": rec.as_row(),
                "next_id": next_tab_id(key),
            })

    if request.method == "DELETE":
        with locks[key]:
            try:
                with transaction.atomic():
                    visible_tab_records(request, key).delete()
            except Exception:
                return JsonResponse({"error": "Failed to clear records."}, status=500)
        return JsonResponse({"ok": True, "next_id": next_tab_id(key)})

    return HttpResponseNotAllowed(["GET", "POST", "DELETE"])


# ---------------------------------------------------------------------------
# Single record: PUT (update) / DELETE
# ---------------------------------------------------------------------------
@api_login_required
def api_record_detail(request, key, rec_id):
    if key not in TAB_ORDER:
        return JsonResponse({"error": "Unknown tab"}, status=404)
    spec = get_spec(key)

    if request.method == "PUT":
        body = parse_json_body(request)
        raw_row = body.get("row")
        if not isinstance(raw_row, dict):
            return JsonResponse({"error": "Malformed request: 'row' must be an object."}, status=400)
        row = {c: str(raw_row.get(c, "") or "") for c in spec["columns"]}
        row.pop("ID", None)
        invalid_dates = invalid_date_fields(spec, row)
        if invalid_dates:
            return JsonResponse(
                {"error": f"{', '.join(invalid_dates)} must use MM-DD-YYYY."},
                status=400,
            )
        coordinate_errors = normalize_coordinate_fields(spec, row)
        if coordinate_errors:
            return JsonResponse({"error": "; ".join(coordinate_errors)}, status=400)

        with locks[key]:
            try:
                rec = visible_tab_records(request, key).get(tab_record_id=rec_id)
            except TabRecord.DoesNotExist:
                return JsonResponse({"error": "Record not found"}, status=404)

            # Image links are generated only by the upload endpoint. Preserve
            # the existing value during normal record edits.
            row["IMAGE_LINK"] = str(rec.data.get("IMAGE_LINK", "") or "")

            ts_dict = get_section_state(key)
            row = compute_auto_fields(key, row, ts_dict)
            missing = missing_required_fields(key, row)
            if missing:
                return JsonResponse({"error": spec["uid_required_msg"]}, status=400)

            try:
                with transaction.atomic():
                    rec.data = row
                    rec.save(update_fields=["data", "updated_at"])
            except Exception:
                return JsonResponse({"error": "Failed to save the record to the database."}, status=500)
            return JsonResponse({"record": rec.as_row()})

    if request.method == "DELETE":
        with locks[key]:
            try:
                rec = visible_tab_records(request, key).get(tab_record_id=rec_id)
            except TabRecord.DoesNotExist:
                return JsonResponse({"error": "Record not found"}, status=404)
            try:
                with transaction.atomic():
                    rec.delete()
            except Exception:
                return JsonResponse({"error": "Failed to delete the record."}, status=500)
            return JsonResponse({"ok": True, "next_id": next_tab_id(key)})

    return HttpResponseNotAllowed(["PUT", "DELETE"])


@api_login_required
@require_POST
def api_record_image(request, key, rec_id):
    if key not in TAB_ORDER:
        return JsonResponse({"error": "Unknown tab"}, status=404)
    try:
        rec = visible_tab_records(request, key).get(tab_record_id=rec_id)
    except TabRecord.DoesNotExist:
        return JsonResponse({"error": "Record not found"}, status=404)
    upload = request.FILES.get("image")
    if upload is None:
        return JsonResponse({"error": "Choose an image to upload."}, status=400)
    try:
        filename, _ = save_record_image(key, rec_id, upload)
    except InventoryImageError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except OSError:
        return JsonResponse({"error": "The image could not be saved on the server."}, status=500)

    folder = next(folder for folder, section in FOLDER_SECTIONS.items() if section == key)
    image_path = f"/uploads/images/{folder}/{rec_id}/{filename}"
    image_url = request.build_absolute_uri(image_path)
    data = dict(rec.data)
    data["IMAGE_LINK"] = image_url
    rec.data = data
    rec.save(update_fields=["data", "updated_at"])
    return JsonResponse({"record": rec.as_row()})


@login_required
@require_GET
def inventory_image(request, folder, rec_id, filename):
    key = FOLDER_SECTIONS.get(folder)
    if key is None or filename not in {"image.jpg", "image.png", "image.webp"}:
        return JsonResponse({"error": "Image not found"}, status=404)
    try:
        visible_tab_records(request, key).get(tab_record_id=rec_id)
    except TabRecord.DoesNotExist:
        return JsonResponse({"error": "Image not found"}, status=404)
    path = record_image_directory(key, rec_id) / filename
    if not path.is_file():
        return JsonResponse({"error": "Image not found"}, status=404)
    content_type = {
        ".jpg": "image/jpeg", ".png": "image/png", ".webp": "image/webp",
    }[path.suffix.lower()]
    response = FileResponse(path.open("rb"), content_type=content_type)
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    response["Cache-Control"] = "private, max-age=3600"
    return response


# ---------------------------------------------------------------------------
# Dropdown "+" add option
# ---------------------------------------------------------------------------
@api_login_required
def api_options(request, key):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    if not request.user.is_staff:
        return JsonResponse(
            {"error": "Administrator permission is required."},
            status=403,
        )
    if key not in TAB_ORDER:
        return JsonResponse({"error": "Unknown tab"}, status=404)
    spec = get_spec(key)

    data = parse_json_body(request)
    field = data.get("field")
    value = str(data.get("value") or "").strip()
    if not field or not isinstance(field, str) or field not in spec["dropdown_fields"]:
        return JsonResponse({"error": "A valid field name is required."}, status=400)
    if not value:
        return JsonResponse({"error": "value is required"}, status=400)
    if len(value) > 500:
        return JsonResponse({"error": "value must be 500 characters or fewer"}, status=400)
    if field not in spec["preserve_case_fields"]:
        value = value.upper()

    with locks[key]:
        DropdownOption.objects.get_or_create(
            section_id=key,
            field_name=field,
            value=value,
            defaults={
                "sort_order": DropdownOption.objects.filter(
                    section_id=key, field_name=field
                ).count()
            },
        )
        state = get_section_state(key)
        result = {"field": field, "value": value, "options": state["options"].get(field, [])}

        ml = spec.get("mutcd_link")
        if ml and field == ml["code_field"] and value not in state["mutcd_to_class"]:
            classification = str(data.get("classification") or "").strip()
            MutcdClassification.objects.get_or_create(
                section_id=key,
                code=value,
                defaults={"classification": classification},
            )
            result["mutcd_to_class_entry"] = {"code": value, "classification": classification}

        elif ml and field == ml["word_field"] and value not in state["mutcd_map"]:
            mutcd_code = str(data.get("mutcd_code") or "").strip()
            classification = state["mutcd_to_class"].get(mutcd_code, "")
            if mutcd_code and not classification:
                classification = str(data.get("classification") or "").strip()
                if classification:
                    MutcdClassification.objects.update_or_create(
                        section_id=key,
                        code=mutcd_code,
                        defaults={"classification": classification},
                    )
            MutcdMapping.objects.create(
                section_id=key,
                word_description=value,
                mutcd_code=mutcd_code,
                classification=classification,
            )
            result["mutcd_map_entry"] = {
                "word": value,
                "MUTCD": mutcd_code,
                "MUTCD_CLASSIFICATION": classification,
            }
            refreshed = get_section_state(key)
            result["mutcd_reverse_map"] = refreshed["mutcd_reverse_map"]
            result["mutcd_word_options"] = refreshed["mutcd_word_options"]

        afm = spec.get("auto_fill_map")
        if afm and field == afm["driver_field"] and value not in state["type_map"]:
            entry = {}
            for dep_field in afm["dependent_fields"]:
                entry[dep_field] = str(data.get(f"dep_{dep_field}") or "").strip()
            AutoFillMapping.objects.create(
                section_id=key,
                driver_value=value,
                values=entry,
            )
            result["type_map_entry"] = {"driver_value": value, "fields": entry}

    return JsonResponse(result)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def write_sheet(key, ws, records):
    spec = get_spec(key)
    columns = spec["columns"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(columns)
    for col_idx, _ in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    def is_number_like(val):
        if val is None:
            return True
        s = str(val).strip()
        if s == "":
            return True
        try:
            float(s)
            return True
        except ValueError:
            return False

    numeric_fields = set(spec.get("extra_numeric_fields", []))
    numeric_fields |= {c for c in columns if "LATITUDE" in c or "LONGITUDE" in c}
    for c in columns:
        if c in numeric_fields:
            continue
        has_value = any(str(row.get(c, "")).strip() != "" for row in records)
        if has_value and all(is_number_like(row.get(c)) for row in records):
            numeric_fields.add(c)

    for row in records:
        values = []
        for c in columns:
            val = row.get(c, "")
            if c in numeric_fields and val not in (None, ""):
                try:
                    fval = float(val)
                    val = int(fval) if fval.is_integer() else fval
                except (TypeError, ValueError):
                    pass
            values.append(safe_excel_cell(val))
        ws.append(values)

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row in records:
            val = str(row.get(col_name, ""))
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@api_login_required
@require_GET
def api_export(request, key):
    if key not in TAB_ORDER:
        return JsonResponse({"error": "Unknown tab"}, status=404)
    spec = get_spec(key)
    with locks[key]:
        records = [
            r.as_row()
            for r in visible_tab_records(request, key).order_by("tab_record_id")
        ]
    if not records:
        return JsonResponse({"error": "There are no records to export yet."}, status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = spec["export_sheet_name"]
    write_sheet(key, ws, records)

    filename = f"{spec['export_filename_prefix']}_{date.today().isoformat()}.xlsx"
    return _xlsx_response(wb, filename)


@api_login_required
@require_GET
def api_export_all(request):
    snapshots = {}
    for key in TAB_ORDER:
        with locks[key]:
            snapshots[key] = [
                r.as_row()
                for r in visible_tab_records(request, key).order_by("tab_record_id")
            ]

    wb = Workbook()
    wb.remove(wb.active)
    any_data = False
    for key in TAB_ORDER:
        if not snapshots[key]:
            continue
        any_data = True
        ws = wb.create_sheet(title=get_spec(key)["export_sheet_name"])
        write_sheet(key, ws, snapshots[key])
    if not any_data:
        return JsonResponse({"error": "There are no records in any tab to export yet."}, status=400)

    filename = f"All_Inventories_{date.today().isoformat()}.xlsx"
    return _xlsx_response(wb, filename)
